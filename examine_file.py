from openpyxl import load_workbook

wb = load_workbook("BenefitHistory.xlsx")
print("Sheet names:", wb.sheetnames)

ws_espp = wb["ESPP"]
print("\n--- ESPP Sheet ---")
print("Columns:", [cell.value for cell in ws_espp[1]])
print("First few rows:")
for i, row in enumerate(ws_espp.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f"Row {i}: {row}")

ws_rs = wb["Restricted Stock"]
print("\n--- Restricted Stock Sheet ---")
print("Columns:", [cell.value for cell in ws_rs[1]])
print("First few rows:")
for i, row in enumerate(ws_rs.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f"Row {i}: {row}")
