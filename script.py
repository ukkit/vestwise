__version__ = "2026.0302.1321"
__author__ = "Neeraj Tikku"
__copyright__ = "Copyright 2026, Neeraj Tikku"

import calendar
import configparser
import os
import urllib.request
import warnings
from datetime import date as date_cls
from datetime import datetime

import pandas as pd

warnings.filterwarnings("ignore")

try:
    import yfinance as yf

    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False
    print("Warning: yfinance not installed. Stock price lookup will be unavailable.")
    print("Install with: pip install yfinance")

try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Indian tax rates for foreign/unlisted shares (post-Budget 2024)
LTCG_RATE = 0.125  # 12.5% for holdings >= 24 months
STCG_RATE = 0.30  # User's marginal slab rate (default 30%)
LTCG_HOLDING_MONTHS = 24  # Unlisted/foreign shares threshold

# SBI TTBR (Telegraphic Transfer Buying Rate) for Rule 115 compliant USD-INR conversion
SBI_TTBR_CSV_URL = (
    "https://raw.githubusercontent.com/sahilgupta/sbi-fx-ratekeeper/main/csv_files/SBI_REFERENCE_RATES_USD.csv"
)
SBI_TTBR_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "SBI_REFERENCE_RATES_USD.csv")
SALE_PRICE_OVERRIDES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "sale_price_overrides.csv")


def _load_config():
    """Load overrides from vestwise.ini (if present). Falls back to hardcoded defaults."""
    global LTCG_RATE, STCG_RATE, LTCG_HOLDING_MONTHS
    global SBI_TTBR_CSV_URL, SBI_TTBR_CACHE_FILE, SALE_PRICE_OVERRIDES_FILE

    config = configparser.ConfigParser()
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vestwise.ini")
    if not config.read(config_path):
        return  # No ini file — silently use hardcoded defaults

    LTCG_RATE = config.getfloat("tax", "ltcg_rate", fallback=LTCG_RATE)
    STCG_RATE = config.getfloat("tax", "stcg_rate", fallback=STCG_RATE)
    LTCG_HOLDING_MONTHS = config.getint("tax", "ltcg_holding_months", fallback=LTCG_HOLDING_MONTHS)
    SBI_TTBR_CSV_URL = config.get("paths", "sbi_ttbr_csv_url", fallback=SBI_TTBR_CSV_URL)

    # Resolve relative paths against script directory
    _base = os.path.dirname(os.path.abspath(__file__))
    raw_cache = config.get("paths", "sbi_ttbr_cache_file", fallback=None)
    if raw_cache:
        SBI_TTBR_CACHE_FILE = os.path.join(_base, raw_cache) if not os.path.isabs(raw_cache) else raw_cache

    raw_overrides = config.get("paths", "sale_price_overrides_file", fallback=None)
    if raw_overrides:
        SALE_PRICE_OVERRIDES_FILE = os.path.join(_base, raw_overrides) if not os.path.isabs(raw_overrides) else raw_overrides


_load_config()


def parse_percentage(value_str):
    """Parse percentage string (e.g., '30.9%') to float."""
    if pd.isna(value_str) or str(value_str).strip() == "":
        return 0

    value_str = str(value_str).strip()
    # Remove percentage sign if present
    value_str = value_str.rstrip("%")

    try:
        return float(value_str)
    except ValueError:
        print(f"Warning: Could not parse percentage value: {value_str}")
        return 0


def parse_date(date_str):
    """Parse date string in various formats."""
    if pd.isna(date_str) or str(date_str).strip() == "":
        return None

    date_str = str(date_str).strip()

    # Try different date formats
    date_formats = [
        "%d-%b-%Y",  # 19-NOV-2025
        "%m/%d/%Y",  # 11/19/2025
        "%d/%m/%Y",  # 19/11/2025 (if needed)
        "%Y-%m-%d",  # 2025-11-19
        "%b %d, %Y",  # Nov 19, 2025
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue

    # If all formats fail, try to extract date parts
    try:
        # Handle cases like "11/19/2025 00:00:00"
        if " " in date_str:
            date_part = date_str.split(" ")[0]
            for fmt in ["%m/%d/%Y", "%Y-%m-%d"]:
                try:
                    return datetime.strptime(date_part, fmt)
                except Exception:
                    continue
    except Exception:
        pass

    print(f"Warning: Could not parse date: {date_str}")
    return None


def get_financial_year(date_obj):
    """
    Get financial year (April 1 - March 31) for a given date.
    For example, April 1, 2025 - March 31, 2026 is FY 2025-26.
    """
    if date_obj is None:
        return None

    if date_obj.month >= 4:  # April onwards
        return f"FY{date_obj.year}-{date_obj.year + 1}"
    else:  # January to March
        return f"FY{date_obj.year - 1}-{date_obj.year}"


def get_capital_gains_tax_rate(acquisition_date, sale_date):
    """
    Calculate capital gains tax rate based on holding period.
    For foreign/unlisted shares (Indian tax rules):
    LTCG (Long-Term): 12.5% for holdings >= 24 months
    STCG (Short-Term): 30% (slab rate) for holdings < 24 months
    Returns (rate, holding_days, tax_type)
    """
    if acquisition_date is None or sale_date is None:
        return None, None

    # Exact calendar month calculation for LTCG threshold
    months_held = (sale_date.year - acquisition_date.year) * 12 + (sale_date.month - acquisition_date.month)
    if sale_date.day < acquisition_date.day:
        months_held -= 1

    if months_held >= LTCG_HOLDING_MONTHS:
        return LTCG_RATE, "LTCG"
    else:
        return STCG_RATE, "STCG"


def _get_sales_history_col_map(writer):
    """
    Read the header row of the already-written 'Sales History' sheet and return
    a mapping of column header names to Excel column letters.
    Returns {} if the sheet doesn't exist or openpyxl is unavailable.
    """
    if not OPENPYXL_AVAILABLE:
        return {}
    try:
        ws = writer.sheets["Sales History"]
    except KeyError:
        return {}
    col_map = {}
    for cell in ws[1]:
        if cell.value is not None:
            col_map[cell.value] = get_column_letter(cell.column)
    return col_map


def _build_tax_summary_formulas(writer, year_tax_df, col_config, sales_col_map, data_row_mapping=None):
    """
    Overlay SUMIFS formulas onto the Amount ($) column of the Year-wise Tax Summary
    sheet for capital-gains rows.

    Parameters
    ----------
    writer : pd.ExcelWriter
    year_tax_df : DataFrame  (full df including helper columns)
    col_config : dict
        - 'fy_col': header name for the FY column in year_tax_df (e.g. 'FY' or 'Financial Year')
    sales_col_map : dict  {header_name: column_letter} from _get_sales_history_col_map
    data_row_mapping : list of (excel_row, df_row_series), optional
        Explicit mapping of Excel row numbers to DataFrame rows.
        If None, assumes consecutive rows starting at row 2.
    """
    if not OPENPYXL_AVAILABLE or not sales_col_map:
        return
    try:
        ws = writer.sheets["Year-wise Tax Summary"]
    except KeyError:
        return

    # Determine Amount ($) column index in the Tax Summary sheet
    ts_col_indices = {cell.value: cell.column for cell in ws[1]}
    amount_col_idx = ts_col_indices.get("Amount ($)")
    if amount_col_idx is None:
        return
    amount_col_letter = get_column_letter(amount_col_idx)

    # Look up Sales History column letters dynamically
    cg_tax_col = sales_col_map.get("Capital Gains Tax ($)")
    grant_id_col = sales_col_map.get("Grant ID")
    symbol_col = sales_col_map.get("Symbol")
    tax_type_col = sales_col_map.get("Tax Type")
    sale_date_col = sales_col_map.get("Sale Date")

    if not all([cg_tax_col, grant_id_col, symbol_col, tax_type_col, sale_date_col]):
        return  # Missing required columns — keep static values

    fy_header = col_config.get("fy_col", "FY")

    # Build iteration source
    if data_row_mapping is not None:
        row_iter = data_row_mapping  # list of (excel_row, row_series)
    else:
        row_iter = [(row_idx, row) for row_idx, (_, row) in enumerate(year_tax_df.iterrows(), start=2)]

    for row_idx, row in row_iter:
        # Skip non-capital-gains rows (withholding tax)
        if not row.get("_is_capital_gains", True):
            continue

        grant_id = str(row["Grant ID"]).replace('"', '""')
        symbol = str(row["Symbol"]).replace('"', '""')
        fy = row[fy_header]

        # Extract the tax type base (LTCG or STCG)
        tax_type_base = row.get("_tax_type_base", None)
        if not tax_type_base:
            continue

        # Derive fiscal year start (YYYY) from FY string like 'FY2025-2026'
        try:
            fy_start_year = int(str(fy)[2:6])
        except Exception:
            continue

        formula = (
            f"=SUMIFS('Sales History'!${cg_tax_col}:${cg_tax_col},"
            f"'Sales History'!${grant_id_col}:${grant_id_col},\"{grant_id}\","
            f"'Sales History'!${symbol_col}:${symbol_col},\"{symbol}\","
            f"'Sales History'!${tax_type_col}:${tax_type_col},\"{tax_type_base}\","
            f"'Sales History'!${sale_date_col}:${sale_date_col},\">=\"&DATE({fy_start_year},4,1),"
            f"'Sales History'!${sale_date_col}:${sale_date_col},\"<=\"&DATE({fy_start_year + 1},3,31))"
        )

        ws[f"{amount_col_letter}{row_idx}"] = formula
        ws[f"{amount_col_letter}{row_idx}"].number_format = "$#,##0.00"


def _write_tax_summary_with_subtotals(writer, year_tax_df, fy_col, display_cols):
    """
    Write Year-wise Tax Summary sheet with FY subtotal rows.

    Returns
    -------
    data_row_mapping : list of (excel_row, row_series)
        Maps each data row's Excel row number to its DataFrame row.
    subtotal_rows : set of int
        Excel row numbers that are subtotal rows (for skip_rows in _format_worksheet).
    """
    from openpyxl.utils import get_column_letter as _gcl

    wb = writer.book
    ws = wb.create_sheet("Year-wise Tax Summary")
    writer.sheets["Year-wise Tax Summary"] = ws

    # Write header
    for col_idx, col_name in enumerate(display_cols, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Build column index lookup
    col_lookup = {name: idx for idx, name in enumerate(display_cols, 1)}
    amt_usd_col_idx = col_lookup.get("Amount ($)")
    amt_inr_col_idx = col_lookup.get("Amount (INR)")
    fy_col_idx = col_lookup.get(fy_col)

    # Group by FY (preserving sort order)
    grouped = year_tax_df.groupby(fy_col, sort=False)

    excel_row = 2
    data_row_mapping = []
    subtotal_rows = set()

    for fy_val, group_df in grouped:
        group_start_row = excel_row
        for _, row in group_df.iterrows():
            for col_idx, col_name in enumerate(display_cols, 1):
                val = row.get(col_name)
                if pd.notna(val) if not isinstance(val, str) else val:
                    ws.cell(row=excel_row, column=col_idx, value=val)
            data_row_mapping.append((excel_row, row))
            excel_row += 1
        group_end_row = excel_row - 1

        # --- Subtotal row ---
        subtotal_row = excel_row
        subtotal_rows.add(subtotal_row)

        if fy_col_idx:
            ws.cell(row=subtotal_row, column=fy_col_idx, value=f"{fy_val} Total")

        if amt_usd_col_idx:
            col_letter = _gcl(amt_usd_col_idx)
            ws.cell(
                row=subtotal_row,
                column=amt_usd_col_idx,
                value=f"=SUM({col_letter}{group_start_row}:{col_letter}{group_end_row})",
            )
            ws.cell(row=subtotal_row, column=amt_usd_col_idx).number_format = "$#,##0.00"

        if amt_inr_col_idx:
            col_letter = _gcl(amt_inr_col_idx)
            ws.cell(
                row=subtotal_row,
                column=amt_inr_col_idx,
                value=f"=SUM({col_letter}{group_start_row}:{col_letter}{group_end_row})",
            )
            ws.cell(row=subtotal_row, column=amt_inr_col_idx).number_format = "#,##0.00"

        # Style the subtotal row
        for col_idx in range(1, len(display_cols) + 1):
            cell = ws.cell(row=subtotal_row, column=col_idx)
            cell.font = _SUBTOTAL_FONT
            cell.fill = _SUBTOTAL_FILL
            cell.border = _THIN_BORDER

        excel_row += 1

    return data_row_mapping, subtotal_rows


# Column header patterns for center-alignment
_CENTER_ALIGN_HEADERS = {
    "Symbol",
    "Tax Type",
    "FY",
    "Financial Year",
    "Grant Type",
    "Is Future",
    "Is Future Vesting",
    "Validation Status",
    "# of Sales",
    "# of Vest Schedules",
    "# of Tax Withholdings",
    "Vest Period",
    "Holding Days",
    "Days to Vesting",
    "Qty. Sold",
    "Quantity Sold",
    "Vested Qty.",
    "Vested Quantity",
    "Released Qty",
    "Released Quantity",
    "Units",
    "Vested to Date",
    "Withheld for Taxes",
    "Released to Account",
    "Sold",
    "Sellable",
    "Calc Sellable",
    "Unvested",
    "Calc Unvested",
    "Future Vesting (from schedules)",
    "Future Vesting Qty",
    "Grant ID",
    "Holding Period",
    "Rate (%)",
    # Schedule FA (Table A3) columns
    "CY",
    "AY",
    "Country Code",
    "Nature of Interest",
    "Date Since Held",
    "Vested in CY",
    "Sold in CY",
    "Shares Held (Dec 31)",
}

# Styles (defined once, reused across all sheets)
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_FONT = Font(name="Calibri", size=10)
_ALT_ROW_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_SUBTOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True)
_OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_OK_FONT = Font(name="Calibri", size=10, color="375623")
_ISSUE_FONT = Font(name="Calibri", size=10, color="C00000", bold=True)


def _format_worksheet(ws, skip_rows=None):
    """
    Apply professional formatting to a worksheet.

    Parameters
    ----------
    ws : openpyxl Worksheet
    skip_rows : set[int], optional
        Row numbers (1-based) to skip (e.g. subtotal rows that have their own styling).
    """
    if not OPENPYXL_AVAILABLE:
        return

    skip_rows = skip_rows or set()

    # Freeze header row
    ws.freeze_panes = "A2"

    # Build header name -> column letter mapping
    header_names = {}
    for cell in ws[1]:
        if cell.value is not None:
            header_names[cell.column] = str(cell.value)

    max_col = ws.max_column
    max_row = ws.max_row

    # --- Header row styling ---
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # --- Data row styling ---
    for row_idx in range(2, max_row + 1):
        if row_idx in skip_rows:
            continue
        is_even = row_idx % 2 == 0
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            if is_even:
                cell.fill = _ALT_ROW_FILL

            # Alignment based on header name
            hdr = header_names.get(col_idx, "")
            if hdr in _CENTER_ALIGN_HEADERS:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif hdr.endswith(("($)", "(INR)", "(%)", "(USD-INR)")):
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # --- Validation Status conditional coloring ---
    vs_col = None
    for col_idx, hdr in header_names.items():
        if hdr == "Validation Status":
            vs_col = col_idx
            break
    if vs_col:
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=vs_col)
            val = str(cell.value).strip() if cell.value else ""
            if val == "OK":
                cell.fill = _OK_FILL
                cell.font = _OK_FONT
            elif val and val != "None":
                cell.font = _ISSUE_FONT

    # --- Auto-fit column widths ---
    for col_idx in range(1, max_col + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # --- Currency / number formats on data rows ---
    for col_idx, hdr in header_names.items():
        if "($)" in hdr or hdr == "Estimated Market Value ($)":
            fmt = "$#,##0.00"
        elif "(INR)" in hdr:
            fmt = "#,##0.00"
        elif hdr in ("Tax Rate (%)", "Rate (%)"):
            fmt = "0.00"
        elif hdr == "Withholding Amount ($)" or hdr == "Tax Withheld ($)":
            fmt = "$#,##0.00"
        else:
            continue
        for row_idx in range(2, max_row + 1):
            if row_idx not in skip_rows:
                ws.cell(row=row_idx, column=col_idx).number_format = fmt


_sbi_ttbr_df = None  # Module-level cache for SBI TTBR data


def _load_sbi_ttbr_data():
    """
    Load SBI TTBR data from cached CSV or download from GitHub.
    Returns a DataFrame indexed by date with 'TT BUY' values, or None on failure.
    """
    global _sbi_ttbr_df
    if _sbi_ttbr_df is not None:
        return _sbi_ttbr_df if not _sbi_ttbr_df.empty else None

    cache_file = SBI_TTBR_CACHE_FILE
    need_download = True

    # Check if cached file exists and is fresh (< 7 days old)
    if os.path.exists(cache_file):
        age_days = (datetime.now().timestamp() - os.path.getmtime(cache_file)) / 86400
        if age_days < 7:
            need_download = False

    if need_download:
        try:
            os.makedirs(os.path.dirname(cache_file), exist_ok=True)
            print("Downloading SBI TTBR rates from GitHub...")
            with urllib.request.urlopen(SBI_TTBR_CSV_URL, timeout=15) as response:
                with open(cache_file, "wb") as f:
                    f.write(response.read())
            print(f"[OK] SBI TTBR rates cached to {cache_file}")
        except Exception as e:
            print(f"[WARNING] Failed to download SBI TTBR data: {e}")
            if not os.path.exists(cache_file):
                _sbi_ttbr_df = pd.DataFrame()  # sentinel to avoid retrying
                return None
            # Use stale cache if download fails

    try:
        df = pd.read_csv(cache_file)
        required_cols = {"DATE", "TT BUY"}
        missing = required_cols - set(df.columns.str.strip())
        if missing:
            print(f"[WARNING] SBI TTBR CSV missing expected columns: {missing}")
            _sbi_ttbr_df = pd.DataFrame()
            return None
        df.columns = df.columns.str.strip()
        df["DATE"] = pd.to_datetime(df["DATE"], format="mixed")
        df["TT BUY"] = pd.to_numeric(df["TT BUY"], errors="coerce")
        # Filter out holidays (TT BUY = 0) and NaN
        df = df[df["TT BUY"].notna() & (df["TT BUY"] > 0)]
        df["date_only"] = df["DATE"].dt.date
        df = df.set_index("date_only").sort_index()
        _sbi_ttbr_df = df
        return df
    except Exception as e:
        print(f"[WARNING] Failed to parse SBI TTBR CSV: {e}")
        _sbi_ttbr_df = pd.DataFrame()  # sentinel to avoid retrying
        return None


# ---------------------------------------------------------------------------
# Sale price overrides – persists actual execution prices across runs
# ---------------------------------------------------------------------------

_sale_price_overrides = None  # Module-level cache (None = not yet loaded)


def load_sale_price_overrides():
    """
    Load sale_price_overrides.csv into a dict keyed by (grant_id, sale_date_iso, sale_seq).
    Returns {} if the file does not exist yet.
    """
    global _sale_price_overrides
    if _sale_price_overrides is not None:
        return _sale_price_overrides

    if not os.path.exists(SALE_PRICE_OVERRIDES_FILE):
        _sale_price_overrides = {}
        return _sale_price_overrides

    try:
        df = pd.read_csv(SALE_PRICE_OVERRIDES_FILE, dtype={"grant_id": str, "sale_seq": int})
        result = {}
        for _, r in df.iterrows():
            key = (str(r["grant_id"]), str(r["sale_date"]), int(r["sale_seq"]))
            result[key] = {
                "sale_price_usd": float(r["sale_price_usd"]),
                "sale_quantity": float(r["sale_quantity"]) if pd.notna(r.get("sale_quantity")) else None,
                "source": str(r.get("source", "xlsx")),
                "notes": str(r.get("notes", "")) if pd.notna(r.get("notes")) else "",
            }
        _sale_price_overrides = result
        print(f"[OK] Loaded {len(result)} sale price overrides from {SALE_PRICE_OVERRIDES_FILE}")
    except Exception as e:
        print(f"[WARNING] Failed to parse sale_price_overrides.csv: {e}")
        _sale_price_overrides = {}

    return _sale_price_overrides


def save_sale_price_overrides(overrides_dict):
    """
    Write the full overrides dict to CSV. Creates data/ dir if needed.
    Sorts by sale_date descending (newest first), then grant_id, sale_seq.
    sale_price_usd is rounded to 2 decimal places.
    """
    os.makedirs(os.path.dirname(SALE_PRICE_OVERRIDES_FILE), exist_ok=True)
    rows = []
    for (grant_id, sale_date, sale_seq), vals in overrides_dict.items():
        rows.append(
            {
                "grant_id": grant_id,
                "sale_date": sale_date,
                "sale_seq": sale_seq,
                "sale_price_usd": round(vals["sale_price_usd"], 2),
                "sale_quantity": vals.get("sale_quantity"),
                "source": vals["source"],
                "notes": vals["notes"],
            }
        )
    rows.sort(key=lambda r: (r["sale_date"], r["grant_id"], r["sale_seq"]), reverse=True)
    df = pd.DataFrame(
        rows, columns=["grant_id", "sale_date", "sale_seq", "sale_price_usd", "sale_quantity", "source", "notes"]
    )
    df.to_csv(SALE_PRICE_OVERRIDES_FILE, index=False)
    print(f"[OK] Saved {len(rows)} sale price overrides to {SALE_PRICE_OVERRIDES_FILE}")


def resolve_sale_price(grant_id, event_date_str, row, symbol_for_price, overrides, sale_seq=1):
    """
    Resolve sale price using priority order:
      1. override file (any existing entry – never overwrite)
      2. xlsx Sale Price column (new entry → append)
      3. yfinance closing price (new entry → append)

    Returns (price, source_tag, should_write_to_overrides).
    should_write_to_overrides is True only when the key is absent from the file.
    """
    parsed = parse_date(event_date_str)
    sale_date_iso = parsed.strftime("%Y-%m-%d") if parsed else event_date_str
    key = (str(grant_id), sale_date_iso, sale_seq)

    # Priority 1: existing entry in override file
    if key in overrides:
        entry = overrides[key]
        return entry["sale_price_usd"], entry["source"], False

    # Priority 2: xlsx Sale Price column
    xlsx_price = None
    raw = row.get("Sale Price")
    if raw is not None and pd.notna(raw):
        try:
            xlsx_price = float(raw)
        except (ValueError, TypeError):
            pass
    if xlsx_price is not None:
        return xlsx_price, "xlsx", True

    # Priority 3: yfinance
    if YFINANCE_AVAILABLE:
        price = get_stock_price(symbol_for_price, event_date_str)
        if price is not None:
            return price, "yfinance", True

    return None, None, False


def get_sbi_ttbr(transaction_date_str):
    """
    Get SBI TTBR rate per Rule 115: last business day of the month preceding
    the transaction month.

    Parameters
    ----------
    transaction_date_str : str
        Date string of the transaction

    Returns
    -------
    float or None
        The TT BUY rate, or None if unavailable
    """
    df = _load_sbi_ttbr_data()
    if df is None:
        return None

    parsed = parse_date(transaction_date_str)
    if parsed is None:
        return None

    # Compute the preceding month
    if parsed.month == 1:
        prev_year = parsed.year - 1
        prev_month = 12
    else:
        prev_year = parsed.year
        prev_month = parsed.month - 1

    # Filter to rows in the preceding month
    last_day = calendar.monthrange(prev_year, prev_month)[1]
    month_start = date_cls(prev_year, prev_month, 1)
    month_end = date_cls(prev_year, prev_month, last_day)

    mask = (df.index >= month_start) & (df.index <= month_end)
    month_data = df.loc[mask]

    if month_data.empty:
        return None

    # Last available date in that month with non-zero TT BUY
    last_row = month_data.iloc[-1]
    return float(last_row["TT BUY"])


def get_exchange_rate(date_str):
    """
    Get historical USD to INR exchange rate for a given date.
    Uses SBI TTBR (Rule 115 compliant) first, falls back to yfinance market rate.
    """
    # Try SBI TTBR first (Rule 115 compliant)
    sbi_rate = get_sbi_ttbr(date_str)
    if sbi_rate is not None:
        return sbi_rate

    # Fallback to yfinance market rate
    print(f"[WARNING] SBI TTBR not available for {date_str}, using yfinance market rate")

    if not YFINANCE_AVAILABLE:
        return None

    try:
        parsed_date = parse_date(date_str)
        if parsed_date is None:
            return None

        ticker = yf.Ticker("USDINR=X")
        start_date = (parsed_date - pd.Timedelta(days=5)).strftime("%Y-%m-%d")
        end_date = (parsed_date + pd.Timedelta(days=5)).strftime("%Y-%m-%d")

        hist = ticker.history(start=start_date, end=end_date)

        if len(hist) > 0:
            if hist.index.tz is not None:
                hist.index = hist.index.tz_localize(None)
            closest_date = hist.index[hist.index.get_indexer([parsed_date], method="nearest")[0]]
            return hist.loc[closest_date, "Close"]
    except Exception as e:
        print(f"[WARNING] yfinance exchange rate lookup failed for {date_str}: {e}")

    return None


def get_stock_price(symbol, date_str):
    """Get historical stock price for a given symbol and date."""
    if not YFINANCE_AVAILABLE:
        return None

    try:
        # Parse the date
        parsed_date = parse_date(date_str)
        if parsed_date is None:
            return None

        # Add ticker to make it valid (e.g., PTC -> PTC)
        ticker = yf.Ticker(symbol)

        # Get historical data around the date
        start_date = (parsed_date - pd.Timedelta(days=5)).strftime("%Y-%m-%d")
        end_date = (parsed_date + pd.Timedelta(days=5)).strftime("%Y-%m-%d")

        hist = ticker.history(start=start_date, end=end_date)

        if len(hist) > 0:
            # Normalize tz-aware index to naive for comparison
            if hist.index.tz is not None:
                hist.index = hist.index.tz_localize(None)
            # Find the closest trading date to the actual date
            if parsed_date.strftime("%Y-%m-%d") in hist.index.strftime("%Y-%m-%d"):
                # Exact date match
                idx = hist.index[hist.index.strftime("%Y-%m-%d") == parsed_date.strftime("%Y-%m-%d")][0]
                return hist.loc[idx, "Close"]
            else:
                # Use closest date
                closest_date = hist.index[hist.index.get_indexer([parsed_date], method="nearest")[0]]
                return hist.loc[closest_date, "Close"]
    except Exception as e:
        print(f"Could not fetch price for {symbol} on {date_str}: {str(e)}")
        return None


def process_restricted_stock(df, symbol_for_price="PTC", grant_type="RSU", overrides=None):
    """
    Process Restricted Stock data and return grants dictionary.

    Parameters:
    -----------
    df : DataFrame
        Input data
    symbol_for_price : str
        Stock ticker symbol for historical price lookup (default: 'PTC')
    grant_type : str
        Type of grant ('RSU' or 'ESPP')

    Returns:
    --------
    dict : Dictionary of processed grants
    """
    # Standardize column names (strip whitespace)
    df.columns = df.columns.str.strip()

    # Remove completely empty rows
    df = df.dropna(how="all")

    # Reset index for easier processing
    df = df.reset_index(drop=True)

    # Dictionary to store grant information
    grants = {}
    current_grant = None
    grant_counter = 0

    # Load overrides if not provided by caller (standalone call path)
    if overrides is None:
        overrides = load_sale_price_overrides()
    _sale_seq_counter = {}  # {(grant_id, sale_date_iso): next_seq}

    # Process each row
    for idx, row in df.iterrows():
        record_type = str(row["Record Type"]).strip() if pd.notna(row.get("Record Type")) else ""

        # Handle Grant records
        if record_type == "Grant":
            grant_counter += 1
            symbol = str(row["Symbol"]).strip() if pd.notna(row.get("Symbol")) else ""
            grant_date_str = str(row["Grant Date"]).strip() if pd.notna(row.get("Grant Date")) else ""

            # Create unique grant ID - use Grant Number if available, otherwise use date + counter
            grant_number = str(row.get("Grant Number", "")).strip() if pd.notna(row.get("Grant Number")) else ""
            if grant_number:
                grant_id = grant_number
            else:
                grant_id = f"{grant_date_str}_{grant_counter}"

            # Parse grant date
            grant_date = parse_date(grant_date_str)

            # Get grant date stock price for capital gains calculation
            grant_price = None
            if YFINANCE_AVAILABLE and symbol:
                grant_price = get_stock_price(symbol, grant_date_str)

            # Initialize grant dictionary
            current_grant = {
                "grant_id": grant_id,
                "grant_type": grant_type,
                "symbol": symbol,
                "grant_date": grant_date,
                "grant_date_str": grant_date_str,
                "grant_price": grant_price,
                "granted_qty": float(row["Granted Qty."]) if pd.notna(row.get("Granted Qty.")) else 0,
                "withheld_qty": float(row["Withheld Qty."]) if pd.notna(row.get("Withheld Qty.")) else 0,
                "vested_qty": float(row["Vested Qty."]) if pd.notna(row.get("Vested Qty.")) else 0,
                "sellable_qty": float(row["Sellable Qty."]) if pd.notna(row.get("Sellable Qty.")) else 0,
                "unvested_qty": float(row["Unvested Qty."]) if pd.notna(row.get("Unvested Qty.")) else 0,
                "released_qty": float(row["Released Qty"]) if pd.notna(row.get("Released Qty")) else 0,
                "est_market_value": float(row["Est. Market Value"]) if pd.notna(row.get("Est. Market Value")) else 0,
                "events": [],  # List of events (vest, release, sell)
                "vest_tranches": [],  # Per-vest-tranche data for cost basis
                "vest_schedules": [],  # List of vest schedules
                "tax_withholdings": [],  # List of tax withholdings
                "sales": [],  # List of sales
                "capital_gains_tax": [],  # List of capital gains taxes
                "total_tax_withheld": 0,
                "total_capital_gains_tax": 0,
                "total_sold_qty": 0,
                "total_sale_proceeds": 0,
                "sale_dates": [],
                "validation_issues": [],
            }

            grants[grant_id] = current_grant

        # Handle Event records (grant, vest, release, sell)
        elif record_type == "Event" and current_grant is not None:
            event_date_str = str(row["Date"]).strip() if pd.notna(row.get("Date")) else ""
            event_type = str(row["Event Type"]).strip() if pd.notna(row.get("Event Type")) else ""
            qty_or_amount = float(row["Qty. or Amount"]) if pd.notna(row.get("Qty. or Amount")) else 0

            event_date = parse_date(event_date_str)

            event_info = {"date": event_date, "date_str": event_date_str, "type": event_type, "quantity": qty_or_amount}

            current_grant["events"].append(event_info)

            # Track vest tranches for RSU cost basis (vest date = acquisition date)
            if "vested" in event_type.lower():
                vest_price = get_stock_price(symbol_for_price, event_date_str) if YFINANCE_AVAILABLE else None
                current_grant["vest_tranches"].append(
                    {
                        "vest_date": event_date,
                        "vest_date_str": event_date_str,
                        "quantity": qty_or_amount,
                        "vest_price": vest_price,
                    }
                )

            # Track sales separately
            if "sold" in event_type.lower():
                # Determine sequence number for this sale (handles same-day multi-sales)
                _parsed_sale = parse_date(event_date_str)
                _sale_date_iso = _parsed_sale.strftime("%Y-%m-%d") if _parsed_sale else event_date_str
                _seq_key = (current_grant["grant_id"], _sale_date_iso)
                sale_seq = _sale_seq_counter.get(_seq_key, 0) + 1
                _sale_seq_counter[_seq_key] = sale_seq

                sale_price, price_source, _write_new = resolve_sale_price(
                    current_grant["grant_id"], event_date_str, row, symbol_for_price, overrides, sale_seq
                )
                if _write_new:
                    overrides[(current_grant["grant_id"], _sale_date_iso, sale_seq)] = {
                        "sale_price_usd": sale_price,
                        "sale_quantity": qty_or_amount,
                        "source": price_source,
                        "notes": "",
                    }

                # Get exchange rate on sale date
                exchange_rate = None
                if YFINANCE_AVAILABLE:
                    exchange_rate = get_exchange_rate(event_date_str)

                # Match sale to vest tranche (most recent vest before or on sale date)
                matched_vest = None
                for vt in reversed(current_grant["vest_tranches"]):
                    if vt["vest_date"] and event_date and vt["vest_date"] <= event_date:
                        matched_vest = vt
                        break

                # Use vest date/price as acquisition date/cost basis for RSUs
                if matched_vest:
                    acquisition_date = matched_vest["vest_date"]
                    cost_basis_price = matched_vest["vest_price"]
                else:
                    # Fallback to grant date if no vest tranche found
                    acquisition_date = current_grant["grant_date"]
                    cost_basis_price = current_grant["grant_price"]

                # Calculate capital gains tax based on holding period
                capital_gain = 0
                capital_gains_tax = 0
                tax_rate = 0
                tax_type = "N/A"

                if sale_price is not None and cost_basis_price is not None:
                    capital_gain = (sale_price - cost_basis_price) * qty_or_amount

                    # Determine tax rate based on holding period from vest date
                    tax_rate, tax_type = get_capital_gains_tax_rate(acquisition_date, event_date)

                    if tax_rate is not None:
                        capital_gains_tax = capital_gain * tax_rate
                        current_grant["total_capital_gains_tax"] += capital_gains_tax

                        holding_days = (event_date - acquisition_date).days if acquisition_date else 0

                        # Track capital gain tax separately
                        current_grant["capital_gains_tax"].append(
                            {
                                "date": event_date,
                                "date_str": event_date_str,
                                "grant_price": cost_basis_price,
                                "sale_price": sale_price,
                                "quantity": qty_or_amount,
                                "capital_gain": capital_gain,
                                "holding_days": holding_days,
                                "tax_type": tax_type,
                                "tax_rate": tax_rate,
                                "tax_amount": capital_gains_tax,
                            }
                        )

                holding_days = (event_date - acquisition_date).days if acquisition_date and event_date else 0

                sale_info = {
                    "date": event_date,
                    "date_str": event_date_str,
                    "quantity": qty_or_amount,
                    "price": sale_price,
                    "price_source": price_source,
                    "grant_price": cost_basis_price,  # FMV on vest date
                    "acquisition_date": acquisition_date,  # Vest date for holding period
                    "capital_gain": capital_gain,
                    "capital_gains_tax": capital_gains_tax,
                    "holding_days": holding_days,
                    "tax_type": tax_type,
                    "tax_rate": tax_rate,
                    "exchange_rate": exchange_rate,
                }
                current_grant["sales"].append(sale_info)
                current_grant["total_sold_qty"] += qty_or_amount
                current_grant["sale_dates"].append(event_date_str)

        # Handle Vest Schedule records
        elif record_type == "Vest Schedule" and current_grant is not None:
            vest_date_str = str(row["Vest Date"]).strip() if pd.notna(row.get("Vest Date")) else ""
            vested_qty = float(row["Vested Qty."]) if pd.notna(row.get("Vested Qty.")) else 0
            released_qty = float(row["Released Qty"]) if pd.notna(row.get("Released Qty")) else 0
            vest_period = str(row["Vest Period"]).strip() if pd.notna(row.get("Vest Period")) else ""

            vest_date = parse_date(vest_date_str)

            vest_schedule = {
                "vest_date": vest_date,
                "vest_date_str": vest_date_str,
                "vested_qty": vested_qty,
                "released_qty": released_qty,
                "vest_period": vest_period,
                "is_future": vest_date > datetime.now() if vest_date else False,
            }

            current_grant["vest_schedules"].append(vest_schedule)

        # Handle Tax Withholding records (only for RSU, not ESPP)
        elif record_type == "Tax Withholding" and current_grant is not None and grant_type == "RSU":
            withholding_date_str = str(row["Date"]).strip() if pd.notna(row.get("Date")) else ""
            tax_rate = parse_percentage(row["Effective Tax Rate"]) if pd.notna(row.get("Effective Tax Rate")) else 0
            withholding_amount = float(row["Withholding Amount"]) if pd.notna(row.get("Withholding Amount")) else 0
            tax_description = str(row["Tax Description"]).strip() if pd.notna(row.get("Tax Description")) else ""

            # Only include non-zero tax rate withholdings
            if tax_rate > 0:
                withholding_date = parse_date(withholding_date_str)

                # Get exchange rate on withholding date
                exchange_rate = None
                if YFINANCE_AVAILABLE and withholding_date_str:
                    exchange_rate = get_exchange_rate(withholding_date_str)

                tax_info = {
                    "date": withholding_date,
                    "date_str": withholding_date_str,
                    "tax_rate": tax_rate,
                    "withholding_amount": withholding_amount,
                    "tax_description": tax_description,
                    "exchange_rate": exchange_rate,
                }

                current_grant["tax_withholdings"].append(tax_info)
                current_grant["total_tax_withheld"] += withholding_amount

    return grants


def process_espp(df, symbol_for_price="PTC", overrides=None):
    """
    Process ESPP data and return grants dictionary.
    ESPP is immediately sellable and taxes are paid after selling.

    Parameters:
    -----------
    df : DataFrame
        Input data from ESPP sheet
    symbol_for_price : str
        Stock ticker symbol for historical price lookup (default: 'PTC')

    Returns:
    --------
    dict : Dictionary of processed ESPP grants
    """
    # Standardize column names (strip whitespace)
    df.columns = df.columns.str.strip()

    # Remove completely empty rows
    df = df.dropna(how="all")

    # Reset index for easier processing
    df = df.reset_index(drop=True)

    # Dictionary to store grant information
    grants = {}
    current_grant = None
    grant_counter = 0

    # Load overrides if not provided by caller
    if overrides is None:
        overrides = load_sale_price_overrides()
    _sale_seq_counter = {}  # {(grant_id, sale_date_iso): next_seq}

    # Process each row
    for idx, row in df.iterrows():
        record_type = str(row["Record Type"]).strip() if pd.notna(row.get("Record Type")) else ""

        # Handle Grant records (for ESPP, this is a Purchase)
        if record_type in ("Grant", "Purchase"):
            grant_counter += 1
            symbol = str(row["Symbol"]).strip() if pd.notna(row.get("Symbol")) else ""
            purchase_date_str = str(row["Purchase Date"]).strip() if pd.notna(row.get("Purchase Date")) else ""

            # Create unique grant ID
            grant_id = f"ESPP_{purchase_date_str}_{grant_counter}"

            # Parse purchase date
            purchase_date = parse_date(purchase_date_str)

            # Get purchase price for capital gains calculation
            purchase_price = float(row["Purchase Price"]) if pd.notna(row.get("Purchase Price")) else None

            # Get grant date (for reference)
            grant_date_str = str(row["Grant Date"]).strip() if pd.notna(row.get("Grant Date")) else purchase_date_str
            grant_date = parse_date(grant_date_str)

            # Get purchased quantity and tax collection shares
            purchased_qty = float(row["Purchased Qty."]) if pd.notna(row.get("Purchased Qty.")) else 0
            tax_collection_shares = (
                float(row["Tax Collection Shares"]) if pd.notna(row.get("Tax Collection Shares")) else 0
            )
            net_shares = float(row["Net Shares"]) if pd.notna(row.get("Net Shares")) else 0
            sellable_qty = float(row["Sellable Qty."]) if pd.notna(row.get("Sellable Qty.")) else 0

            # Initialize ESPP grant dictionary
            current_grant = {
                "grant_id": grant_id,
                "grant_type": "ESPP",
                "symbol": symbol,
                "grant_date": grant_date,
                "grant_date_str": grant_date_str,
                "purchase_date": purchase_date,
                "purchase_date_str": purchase_date_str,
                "grant_price": purchase_price,  # Use purchase price as basis
                "granted_qty": purchased_qty,  # Use purchased qty
                "withheld_qty": tax_collection_shares,
                "vested_qty": net_shares,  # All purchased shares are immediately available
                "sellable_qty": sellable_qty,
                "unvested_qty": 0,  # ESPP is immediately vested/sellable
                "released_qty": 0,
                "est_market_value": float(row["Est. Market Value"]) if pd.notna(row.get("Est. Market Value")) else 0,
                "events": [],  # List of events (sell, dividend, etc)
                "vest_schedules": [],  # Not applicable for ESPP
                "tax_withholdings": [],  # Not applicable for ESPP (taxes paid after sale)
                "sales": [],  # List of sales
                "capital_gains_tax": [],  # List of capital gains taxes
                "total_tax_withheld": 0,  # Will be calculated from sales tax
                "total_capital_gains_tax": 0,
                "total_sold_qty": 0,
                "total_sale_proceeds": 0,
                "sale_dates": [],
                "validation_issues": [],
            }

            grants[grant_id] = current_grant

        # Handle Event records
        elif record_type == "Event" and current_grant is not None:
            event_date_str = str(row["Date"]).strip() if pd.notna(row.get("Date")) else ""
            event_type = str(row["Event Type"]).strip() if pd.notna(row.get("Event Type")) else ""
            qty = float(row["Qty"]) if pd.notna(row.get("Qty")) else 0

            event_date = parse_date(event_date_str)

            event_info = {"date": event_date, "date_str": event_date_str, "type": event_type, "quantity": qty}

            current_grant["events"].append(event_info)

            # Track sales
            if "sold" in event_type.lower():
                # Determine sequence number for this sale (handles same-day multi-sales)
                _parsed_sale = parse_date(event_date_str)
                _sale_date_iso = _parsed_sale.strftime("%Y-%m-%d") if _parsed_sale else event_date_str
                _seq_key = (current_grant["grant_id"], _sale_date_iso)
                sale_seq = _sale_seq_counter.get(_seq_key, 0) + 1
                _sale_seq_counter[_seq_key] = sale_seq

                sale_price, price_source, _write_new = resolve_sale_price(
                    current_grant["grant_id"], event_date_str, row, symbol_for_price, overrides, sale_seq
                )
                if _write_new:
                    overrides[(current_grant["grant_id"], _sale_date_iso, sale_seq)] = {
                        "sale_price_usd": sale_price,
                        "sale_quantity": qty,
                        "source": price_source,
                        "notes": "",
                    }

                # Get exchange rate on sale date
                exchange_rate = None
                if YFINANCE_AVAILABLE:
                    exchange_rate = get_exchange_rate(event_date_str)

                # Calculate capital gains tax based on holding period
                capital_gain = 0
                capital_gains_tax = 0
                tax_rate = 0
                tax_type = "N/A"

                if sale_price is not None and current_grant["grant_price"] is not None:
                    capital_gain = (sale_price - current_grant["grant_price"]) * qty

                    # Determine tax rate based on holding period
                    tax_rate, tax_type = get_capital_gains_tax_rate(current_grant["purchase_date"], event_date)

                    if tax_rate is not None:
                        capital_gains_tax = capital_gain * tax_rate
                        current_grant["total_capital_gains_tax"] += capital_gains_tax

                        # Track capital gain tax separately
                        current_grant["capital_gains_tax"].append(
                            {
                                "date": event_date,
                                "date_str": event_date_str,
                                "grant_price": current_grant["grant_price"],
                                "sale_price": sale_price,
                                "quantity": qty,
                                "capital_gain": capital_gain,
                                "holding_days": (event_date - current_grant["purchase_date"]).days,
                                "tax_type": tax_type,
                                "tax_rate": tax_rate,
                                "tax_amount": capital_gains_tax,
                            }
                        )

                sale_info = {
                    "date": event_date,
                    "date_str": event_date_str,
                    "quantity": qty,
                    "price": sale_price,
                    "price_source": price_source,
                    "grant_price": current_grant["grant_price"],
                    "capital_gain": capital_gain,
                    "capital_gains_tax": capital_gains_tax,
                    "tax_type": tax_type,
                    "tax_rate": tax_rate,
                    "exchange_rate": exchange_rate,
                }
                current_grant["sales"].append(sale_info)
                current_grant["total_sold_qty"] += qty
                current_grant["sale_dates"].append(event_date_str)

    return grants


def _fifo_date_since_held(vests_sorted, sales_sorted, cutoff):
    """
    Returns the vest_date_str of the oldest unsold vest tranche as of `cutoff` (datetime).
    FIFO: oldest vests are consumed by oldest sales first.
    Returns None if all shares sold.
    """
    vests = [
        (vt["vest_date"], vt["vest_date_str"], float(vt["quantity"]))
        for vt in vests_sorted
        if vt["vest_date"] and vt["vest_date"] <= cutoff
    ]
    sold_remaining = sum(float(s["quantity"]) for s in sales_sorted if s["date"] and s["date"] <= cutoff)
    for _vest_date, vest_date_str, qty in vests:
        if sold_remaining <= 0:
            return vest_date_str
        sold_remaining -= qty
        if sold_remaining < 0:
            return vest_date_str  # tranche partially consumed, some shares still held
    return None  # all shares sold


def _write_schedule_fa_table_a3(writer, grants):
    """
    Generate Schedule FA Table A3 sheet for ITR filing.

    Table A3: Details of Foreign Equity and Debt Interest held in any entity.
    One row per (Indian Financial Year, company symbol) covering the full vesting history.

    Indian FY: Apr 1 – Mar 31.  Schedule FA uses Calendar Year (Jan–Dec) for closing balances.
    For each FY row, the relevant CY is the FY's start year (e.g. FY2024-25 → CY2024 → Dec 31, 2024).
    """
    # --- Collect per-symbol vests and sales ---
    symbol_data: dict = {}

    for _grant_id, grant in grants.items():
        symbol = grant["symbol"]
        if not symbol:
            continue

        if symbol not in symbol_data:
            symbol_data[symbol] = {"vests": [], "sales": []}

        sd = symbol_data[symbol]

        # Build vest-date → price lookup from vest_tranches (gross vested, but price is valid)
        vest_tranches = grant.get("vest_tranches", [])
        vt_price_by_date = {
            vt["vest_date"].date(): vt["vest_price"]
            for vt in vest_tranches
            if vt.get("vest_date") is not None and vt.get("vest_price") is not None
        }

        # Collect NET shares received using "released" events (RSU) or "PURCHASE" events (ESPP).
        # This avoids double-counting tax-withheld shares that never entered the account.
        for event in grant.get("events", []):
            etype = event.get("type", "").lower()
            if ("released" in etype or etype == "purchase") and event.get("date") is not None:
                vest_price = vt_price_by_date.get(event["date"].date())
                if vest_price is None and etype == "purchase":
                    vest_price = grant.get("grant_price")
                sd["vests"].append(
                    {
                        "vest_date": event["date"],
                        "vest_date_str": event["date_str"],
                        "quantity": float(event["quantity"] or 0),
                        "vest_price": vest_price,
                    }
                )

        for sale in grant["sales"]:
            if sale.get("date") is not None:
                sd["sales"].append(
                    {
                        "date": sale["date"],
                        "quantity": float(sale.get("quantity") or 0),
                        "price": sale.get("price"),
                        "exchange_rate": sale.get("exchange_rate"),
                    }
                )

    if not symbol_data:
        return

    # Sort vests and sales chronologically
    for sd in symbol_data.values():
        sd["vests"].sort(key=lambda v: v["vest_date"])
        sd["sales"].sort(key=lambda s: s["date"])

    # --- Determine calendar years with any activity ---
    cy_years: set = set()
    for sd in symbol_data.values():
        for vt in sd["vests"]:
            cy_years.add(vt["vest_date"].year)
        for s in sd["sales"]:
            cy_years.add(s["date"].year)

    # Always include the current CY
    cy_years.add(datetime.now().year)

    # --- Build one row per (CY, symbol) ---
    fa_rows = []

    for y in sorted(cy_years):
        cy_start_dt = datetime(y, 1, 1)
        cy_end_dt = datetime(y, 12, 31)
        cy_prev_end_dt = datetime(y - 1, 12, 31)
        dec31_date_str = f"{y}-12-31"

        for symbol, sd in symbol_data.items():
            all_vests = sd["vests"]
            all_sales = sd["sales"]

            # Quantities vested / sold within the CY window
            vested_in_cy = sum(vt["quantity"] for vt in all_vests if cy_start_dt <= vt["vest_date"] <= cy_end_dt)
            sold_in_cy = sum(s["quantity"] for s in all_sales if cy_start_dt <= s["date"] <= cy_end_dt)

            # Cumulative holdings as of Dec 31
            cum_vested_dec31 = sum(vt["quantity"] for vt in all_vests if vt["vest_date"] <= cy_end_dt)
            cum_sold_dec31 = sum(s["quantity"] for s in all_sales if s["date"] <= cy_end_dt)
            shares_held_dec31 = cum_vested_dec31 - cum_sold_dec31

            # Skip rows with no activity and no holdings
            if vested_in_cy == 0 and sold_in_cy == 0 and shares_held_dec31 <= 0:
                continue

            # Acquisition value: releases within this CY
            acq_value_usd = 0.0
            acq_value_inr = 0.0
            has_acq_value = False
            for vt in all_vests:
                if cy_start_dt <= vt["vest_date"] <= cy_end_dt and vt["vest_price"] is not None:
                    exch = get_exchange_rate(vt["vest_date_str"]) or 0.0
                    acq_value_usd += vt["vest_price"] * vt["quantity"]
                    acq_value_inr += vt["vest_price"] * vt["quantity"] * exch
                    has_acq_value = True

            # Sale proceeds: sales within this CY
            sale_proceeds_usd = 0.0
            sale_proceeds_inr = 0.0
            has_sale_proceeds = False
            for s in all_sales:
                if cy_start_dt <= s["date"] <= cy_end_dt and s.get("price") is not None:
                    proceeds = s["price"] * s["quantity"]
                    sale_proceeds_usd += proceeds
                    sale_proceeds_inr += proceeds * (s.get("exchange_rate") or 0.0)
                    has_sale_proceeds = True

            # Dec 31 market data (nearest available trading day)
            dec31_price = get_stock_price(symbol, dec31_date_str) if YFINANCE_AVAILABLE else None
            dec31_rate = get_exchange_rate(dec31_date_str) or None

            # Peak price: highest intraday high during the CY
            peak_price_cy = None
            if YFINANCE_AVAILABLE:
                try:
                    hist = yf.Ticker(symbol).history(start=f"{y}-01-01", end=f"{y + 1}-01-01")
                    if not hist.empty:
                        peak_price_cy = float(hist["High"].max())
                except Exception:
                    pass

            # Peak shares = shares at Jan 1 + all releases during CY (before any CY sales)
            cum_vested_prev = sum(vt["quantity"] for vt in all_vests if vt["vest_date"] <= cy_prev_end_dt)
            cum_sold_prev = sum(s["quantity"] for s in all_sales if s["date"] <= cy_prev_end_dt)
            peak_shares = (cum_vested_prev - cum_sold_prev) + vested_in_cy

            peak_balance_inr = (
                peak_shares * peak_price_cy * dec31_rate
                if peak_price_cy is not None and dec31_rate is not None and peak_shares > 0
                else None
            )
            closing_balance_inr = (
                shares_held_dec31 * dec31_price * dec31_rate
                if dec31_price is not None and dec31_rate is not None and shares_held_dec31 > 0
                else None
            )

            # Date Since Held: FIFO cutoff = Dec 31 (matches Schedule FA reporting date)
            date_since_held = _fifo_date_since_held(all_vests, all_sales, cy_end_dt)

            fa_rows.append(
                {
                    "CY": f"CY{y}",
                    "AY": f"AY{y + 1}-{y + 2}",
                    "Country Code": "US",
                    "Name of Entity (Ticker)": symbol,
                    "Nature of Interest": "Direct",
                    "Date Since Held": date_since_held or "",
                    "Vested in CY": vested_in_cy if vested_in_cy > 0 else None,
                    "Sold in CY": sold_in_cy if sold_in_cy > 0 else None,
                    "Shares Held (Dec 31)": shares_held_dec31 if shares_held_dec31 > 0 else (0 if cum_vested_dec31 > 0 else None),
                    "Acquisition Value ($)": acq_value_usd if has_acq_value else None,
                    "Acquisition Value (INR)": acq_value_inr if has_acq_value else None,
                    "Dec 31 Price ($)": dec31_price,
                    "Dec 31 Rate (USD-INR)": dec31_rate,
                    "Peak Balance (INR)": peak_balance_inr,
                    "Closing Balance (INR)": closing_balance_inr,
                    "Sale Proceeds ($)": sale_proceeds_usd if has_sale_proceeds else None,
                    "Sale Proceeds (INR)": sale_proceeds_inr if has_sale_proceeds else None,
                }
            )

    # Sort: most recent CY first, then by symbol
    fa_rows.sort(key=lambda r: (-int(str(r["CY"])[2:]), r["Name of Entity (Ticker)"]))

    if not fa_rows:
        return

    fa_df = pd.DataFrame(fa_rows)
    fa_df.to_excel(writer, sheet_name="Schedule FA (Table A3)", index=False)

    if not OPENPYXL_AVAILABLE:
        return

    ws = writer.sheets["Schedule FA (Table A3)"]
    _format_worksheet(ws)

    # Explicit number formats not handled by _format_worksheet column-suffix logic
    col_indices = {cell.value: cell.column for cell in ws[1]}
    num_data_rows = len(fa_rows)

    for col_name, fmt in [
        ("Dec 31 Rate (USD-INR)", "0.0000"),
        ("Vested in CY", "#,##0"),
        ("Sold in CY", "#,##0"),
        ("Shares Held (Dec 31)", "#,##0"),
    ]:
        col_idx = col_indices.get(col_name)
        if col_idx:
            for row_idx in range(2, num_data_rows + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = fmt

    # Notes section below data
    note_row = num_data_rows + 3
    notes = [
        ("NOTES — Schedule FA Table A3 (ITR-2/ITR-3) | One row per Calendar Year (Jan–Dec) per company", True),
        ("  CY = Calendar Year (Jan 1 – Dec 31). AY = Assessment Year for ITR filing (CY+1 to CY+2).", False),
        ("  'Date Since Held' uses FIFO logic — shows the vest date of the oldest UNSOLD tranche as of Dec 31.", False),
        ("  'Shares Held (Dec 31)' = closing balance to report in Schedule FA.", False),
        ("  'Peak Balance (INR)' = (shares at Jan 1 + released in CY) × peak CY intraday high × Dec 31 SBI TTBR.", False),
        ("  'Closing Balance (INR)' = Shares Held (Dec 31) × Dec 31 Price × Dec 31 SBI TTBR.", False),
        ("  'Acquisition Value' = FMV at release × released qty × SBI TTBR on release date (cost basis of received shares).", False),
        ("  Replace 'Name of Entity (Ticker)' with the full company name and registered address before filing ITR.", False),
        ("  File Form 67 BEFORE submitting ITR if claiming foreign tax credit on dividends.", False),
        ("  Non-disclosure of foreign assets attracts Rs.10 lakh penalty under the Black Money Act.", False),
    ]

    bold_font = Font(name="Calibri", size=9, bold=True, color="2F5496")
    note_font = Font(name="Calibri", size=9, italic=True, color="595959")
    max_col = ws.max_column

    for i, (note_text, is_heading) in enumerate(notes):
        cell = ws.cell(row=note_row + i, column=1)
        cell.value = note_text
        cell.font = bold_font if is_heading else note_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if max_col > 1:
            ws.merge_cells(
                start_row=note_row + i,
                start_column=1,
                end_row=note_row + i,
                end_column=max_col,
            )
        ws.row_dimensions[note_row + i].height = 18

    ws.row_dimensions[note_row].height = 20


def process_benefit_history(input_file, output_file, symbol_for_price="PTC"):
    """
    Process BenefitHistory Excel file with multiple sheets (ESPP and Restricted Stock).
    Generates combined summary.

    Parameters:
    -----------
    input_file : str
        Path to input Excel file
    output_file : str
        Path for output Excel file
    symbol_for_price : str
        Stock ticker symbol for historical price lookup (default: 'PTC')
    """

    print(f"Reading file: {input_file}")

    # Try to read both sheets from BenefitHistory.xlsx
    try:
        espp_df = pd.read_excel(input_file, sheet_name="ESPP")
        print("Found ESPP sheet")
    except Exception:
        espp_df = None
        print("ESPP sheet not found")

    try:
        rs_df = pd.read_excel(input_file, sheet_name="Restricted Stock")
        print("Found Restricted Stock sheet")
    except Exception:
        rs_df = None
        print("Restricted Stock sheet not found")

    # If neither sheet found, try old format (single sheet)
    if espp_df is None and rs_df is None:
        print("BenefitHistory format not found, trying old single-sheet format")
        rs_df = pd.read_excel(input_file)

    # Load sale price overrides once; sub-functions mutate it in-place with new entries
    overrides = load_sale_price_overrides()

    # Process both sheets if available
    all_grants = {}

    if rs_df is not None:
        print("\nProcessing Restricted Stock sheet...")
        rs_grants = process_restricted_stock(rs_df, symbol_for_price, grant_type="RSU", overrides=overrides)
        all_grants.update(rs_grants)
        print(f"Found {len(rs_grants)} Restricted Stock grants")

    if espp_df is not None:
        print("\nProcessing ESPP sheet...")
        espp_grants = process_espp(espp_df, symbol_for_price, overrides=overrides)
        all_grants.update(espp_grants)
        print(f"Found {len(espp_grants)} ESPP grants")

    # Persist any newly discovered sale prices (existing entries were never touched)
    save_sale_price_overrides(overrides)

    grants = all_grants

    print(f"Found {len(grants)} total grants")

    # Process and validate each grant
    summary_data = []

    for grant_id, grant in grants.items():
        # Calculate derived values
        total_released = sum(event["quantity"] for event in grant["events"] if "released" in event["type"].lower())

        # Calculate future vesting from schedules
        future_vesting_qty = sum(
            schedule["vested_qty"] for schedule in grant["vest_schedules"] if schedule["is_future"]
        )

        # Calculate next vest date
        future_vest_dates = [
            schedule["vest_date"]
            for schedule in grant["vest_schedules"]
            if schedule["is_future"] and schedule["vest_date"]
        ]
        next_vest_date = min(future_vest_dates) if future_vest_dates else None

        # Calculate sellable quantity
        # RSU: shares are withheld for tax at vesting, so Released = Vested - tax shares
        # ESPP: all purchased shares are immediately sellable, no release step
        grant_type = grant.get("grant_type", "RSU")
        if grant_type == "ESPP":
            calculated_sellable = grant["vested_qty"] - grant["total_sold_qty"]
        else:
            calculated_sellable = total_released - grant["total_sold_qty"]

        # Calculate unvested quantity (alternative calculation)
        calculated_unvested = grant["granted_qty"] - grant["vested_qty"]

        # Validation checks
        validation_issues = []

        # Check 1: Granted = Vested + Unvested
        if abs(grant["granted_qty"] - (grant["vested_qty"] + grant["unvested_qty"])) > 0.01:
            validation_issues.append(
                f"Granted ({grant['granted_qty']}) ≠ Vested ({grant['vested_qty']}) + Unvested ({grant['unvested_qty']})"
            )

        # Check 2: Sellable Qty matches calculation
        if abs(grant["sellable_qty"] - calculated_sellable) > 0.01:
            validation_issues.append(
                f"Sellable Qty mismatch: Stored={grant['sellable_qty']}, Calculated={calculated_sellable}"
            )

        # Check 3: Unvested Qty matches calculation
        if abs(grant["unvested_qty"] - calculated_unvested) > 0.01:
            validation_issues.append(
                f"Unvested Qty mismatch: Stored={grant['unvested_qty']}, Calculated={calculated_unvested}"
            )

        # Format sale dates
        sale_dates_str = "; ".join(sorted(set(grant["sale_dates"]))) if grant["sale_dates"] else "None"

        # Format next vest date
        next_vest_str = next_vest_date.strftime("%Y-%m-%d") if next_vest_date else "N/A"

        # Format validation issues
        validation_str = " | ".join(validation_issues) if validation_issues else "OK"

        # Prepare summary row with Grant Type
        summary_row = {
            "Grant Type": grant_type,
            "Grant ID": grant["grant_id"],
            "Symbol": grant["symbol"],
            "Grant Date": grant["grant_date_str"],
            "Units": grant["granted_qty"],
            "Vested to Date": grant["vested_qty"],
            "Withheld for Taxes": grant["withheld_qty"],
            "Released to Account": total_released,
            "Tax Withheld ($)": grant["total_tax_withheld"],
            "Sold": grant["total_sold_qty"],
            "Sale Dates": sale_dates_str,
            "Sellable": grant["sellable_qty"],
            "Unvested": grant["unvested_qty"],
            "Future Vesting (from schedules)": future_vesting_qty,
            "Next Vest Date": next_vest_str,
            "Estimated Market Value ($)": grant["est_market_value"],
            "Validation Status": validation_str,
            "# of Sales": len(grant["sales"]),
            "# of Vest Schedules": len(grant["vest_schedules"]),
            "# of Tax Withholdings": len(grant["tax_withholdings"]),
        }

        summary_data.append(summary_row)

    # Create summary DataFrame
    summary_df = pd.DataFrame(summary_data)

    # Sort by Grant Date
    summary_df["Grant Date Parsed"] = pd.to_datetime(summary_df["Grant Date"], errors="coerce")
    summary_df = summary_df.sort_values("Grant Date Parsed", ascending=False)
    summary_df = summary_df.drop("Grant Date Parsed", axis=1)

    # Create additional sheets for detailed views
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Main summary sheet
        summary_df.to_excel(writer, sheet_name="Grant Summary", index=False)

        # Create year-wise tax summary sheet
        year_tax_data = []
        tax_summary_subtotal_rows = set()
        tax_summary_data_row_mapping = []

        # Add capital gains taxes from sales (kept as rows so we can build formulas)
        for grant_id, grant in grants.items():
            for cg_tax in grant["capital_gains_tax"]:
                fy = get_financial_year(cg_tax["date"]) if cg_tax["date"] else get_financial_year(grant["grant_date"])
                # Get exchange rate from the matching sale
                exchange_rate = None
                for sale in grant["sales"]:
                    if sale["date_str"] == cg_tax["date_str"]:
                        exchange_rate = sale["exchange_rate"]
                        break
                tax_row = {
                    "FY": fy,
                    "Grant Type": grant.get("grant_type", "RSU"),
                    "Grant ID": grant["grant_id"],
                    "Symbol": grant["symbol"],
                    "Tax Type": f"Capital Gains ({cg_tax['tax_type']})",
                    "Amount ($)": cg_tax["tax_amount"],
                    "Exchange Rate (USD-INR)": exchange_rate,
                    "Amount (INR)": None,  # Will be formula
                    "_tax_type_base": cg_tax["tax_type"],  # helper for formula matching (LTCG/STCG)
                    "_is_capital_gains": True,
                }
                year_tax_data.append(tax_row)

        if year_tax_data:
            year_tax_df = pd.DataFrame(year_tax_data)
            year_tax_df = year_tax_df.sort_values(["FY", "Grant Type"], ascending=[False, True])

            # Write with FY subtotal rows
            display_cols = [
                "FY",
                "Grant Type",
                "Grant ID",
                "Symbol",
                "Tax Type",
                "Amount ($)",
                "Exchange Rate (USD-INR)",
                "Amount (INR)",
            ]
            tax_summary_data_row_mapping, tax_summary_subtotal_rows = _write_tax_summary_with_subtotals(
                writer, year_tax_df, "FY", display_cols
            )

        # Create detailed vesting schedule sheet
        vesting_data = []
        for grant_id, grant in grants.items():
            # Calculate expected shares per future vest from grant's unvested qty
            num_future_vests = sum(1 for s in grant["vest_schedules"] if s["is_future"])
            expected_per_vest = (grant["unvested_qty"] / num_future_vests) if num_future_vests > 0 else 0

            for schedule in grant["vest_schedules"]:
                vesting_row = {
                    "Grant Type": grant.get("grant_type", "RSU"),
                    "Grant ID": grant["grant_id"],
                    "Symbol": grant["symbol"],
                    "Grant Date": grant["grant_date_str"],
                    "Vest Date": schedule["vest_date"],  # datetime for formulas
                    "Vest Period": schedule["vest_period"],
                    "Vested Qty.": schedule["vested_qty"],
                    "Released Qty": schedule["released_qty"],
                    "Is Future": "Yes" if schedule["is_future"] else "No",
                    "Days to Vesting": None,  # Formula placeholder
                    "Future Vesting Qty": expected_per_vest if schedule["is_future"] else None,
                }
                vesting_data.append(vesting_row)

        if vesting_data:
            vesting_df = pd.DataFrame(vesting_data)
            # Sort by Vest Date descending (future dates first)
            vesting_df = vesting_df.sort_values("Vest Date", ascending=False)
            vesting_df.to_excel(writer, sheet_name="Vesting Schedule", index=False)

            if OPENPYXL_AVAILABLE:
                ws_vest = writer.sheets["Vesting Schedule"]
                vest_col_indices = {cell.value: cell.column for cell in ws_vest[1]}
                vest_date_col_idx = vest_col_indices.get("Vest Date")
                days_col_idx = vest_col_indices.get("Days to Vesting")

                if vest_date_col_idx:
                    vest_date_letter = get_column_letter(vest_date_col_idx)
                    for row_idx in range(2, len(vesting_data) + 2):
                        ws_vest.cell(row=row_idx, column=vest_date_col_idx).number_format = "YYYY-MM-DD"

                if days_col_idx and vest_date_col_idx:
                    days_letter = get_column_letter(days_col_idx)
                    vest_date_letter = get_column_letter(vest_date_col_idx)
                    for row_idx in range(2, len(vesting_data) + 2):
                        ws_vest[f"{days_letter}{row_idx}"] = (
                            f'=IF({vest_date_letter}{row_idx}>TODAY(), {vest_date_letter}{row_idx}-TODAY(), "")'
                        )
                        ws_vest[f"{days_letter}{row_idx}"].number_format = "0"

        # Create sales history sheet
        sales_data = []
        for grant_id, grant in grants.items():
            for sale in grant["sales"]:
                sales_row = {
                    "Grant Type": grant.get("grant_type", "RSU"),
                    "Grant ID": grant["grant_id"],
                    "Symbol": grant["symbol"],
                    "Grant Date": grant["grant_date_str"],
                    "Sale Date": sale["date"],  # datetime for SUMIFS date comparisons
                    "Qty. Sold": sale["quantity"],
                    "Sale Price ($)": sale["price"],
                    "Sale Price Source": sale.get("price_source", ""),
                    "Grant Price ($)": sale["grant_price"],
                    "Capital Gain ($)": sale["capital_gain"],
                    "Holding Days": sale.get("holding_days", (sale["date"] - grant["grant_date"]).days)
                    if sale["date"] and grant["grant_date"]
                    else 0,
                    "Tax Type": sale["tax_type"],
                    "Tax Rate (%)": sale["tax_rate"] * 100 if sale["tax_rate"] else 0,
                    "Capital Gains Tax ($)": sale["capital_gains_tax"],
                    "Exchange Rate (USD-INR)": sale["exchange_rate"],
                    "Capital Gain (INR)": None,  # Will be calculated by formula
                    "Capital Gains Tax (INR)": None,  # Will be calculated by formula
                }
                sales_data.append(sales_row)

        if sales_data:
            sales_df = pd.DataFrame(sales_data)
            # Sort by Sale Date (latest first) — Sale Date is already datetime
            sales_df = sales_df.sort_values("Sale Date", ascending=False)

            sales_df.to_excel(writer, sheet_name="Sales History", index=False)

            if OPENPYXL_AVAILABLE:
                worksheet = writer.sheets["Sales History"]

                # Format Sale Date column as date
                col_indices = {col: idx for idx, col in enumerate(sales_df.columns, 1)}
                sale_date_col_idx = col_indices.get("Sale Date", None)
                if sale_date_col_idx:
                    for row_idx in range(2, len(sales_data) + 2):
                        worksheet.cell(row=row_idx, column=sale_date_col_idx).number_format = "YYYY-MM-DD"

                # Add formulas for calculated columns
                for row_idx, (idx, row) in enumerate(sales_df.iterrows(), start=2):
                    cap_gain_col = get_column_letter(col_indices.get("Capital Gain ($)", 1))
                    sale_price_col = get_column_letter(col_indices.get("Sale Price ($)", 1))
                    grant_price_col = get_column_letter(col_indices.get("Grant Price ($)", 1))
                    qty_col = get_column_letter(col_indices.get("Qty. Sold", 1))
                    tax_rate_col = get_column_letter(col_indices.get("Tax Rate (%)", 1))
                    cap_gain_tax_col = get_column_letter(col_indices.get("Capital Gains Tax ($)", 1))
                    exchange_col = get_column_letter(col_indices.get("Exchange Rate (USD-INR)", 1))
                    cap_gain_inr_col = get_column_letter(col_indices.get("Capital Gain (INR)", 1))
                    cap_gain_tax_inr_col = get_column_letter(col_indices.get("Capital Gains Tax (INR)", 1))

                    # Capital Gain ($) = (Sale Price - Grant Price) * Quantity
                    worksheet[f"{cap_gain_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({sale_price_col}{row_idx}), ISNUMBER({grant_price_col}{row_idx}), ISNUMBER({qty_col}{row_idx})), ({sale_price_col}{row_idx} - {grant_price_col}{row_idx}) * {qty_col}{row_idx}, "")'
                    )

                    # Capital Gains Tax ($) = Capital Gain * Tax Rate / 100
                    worksheet[f"{cap_gain_tax_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({cap_gain_col}{row_idx}), ISNUMBER({tax_rate_col}{row_idx})), {cap_gain_col}{row_idx} * {tax_rate_col}{row_idx} / 100, "")'
                    )

                    # Capital Gain (INR) = Capital Gain ($) * Exchange Rate
                    worksheet[f"{cap_gain_inr_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({cap_gain_col}{row_idx}), ISNUMBER({exchange_col}{row_idx}), {exchange_col}{row_idx}<>0), {cap_gain_col}{row_idx} * {exchange_col}{row_idx}, "")'
                    )

                    # Capital Gains Tax (INR) = Capital Gains Tax ($) * Exchange Rate
                    worksheet[f"{cap_gain_tax_inr_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({cap_gain_tax_col}{row_idx}), ISNUMBER({exchange_col}{row_idx}), {exchange_col}{row_idx}<>0), {cap_gain_tax_col}{row_idx} * {exchange_col}{row_idx}, "")'
                    )

                    # Apply currency formatting
                    for col in [cap_gain_col, sale_price_col, grant_price_col, cap_gain_tax_col]:
                        worksheet[f"{col}{row_idx}"].number_format = "$#,##0.00"
                    for col in [cap_gain_inr_col, cap_gain_tax_inr_col]:
                        worksheet[f"{col}{row_idx}"].number_format = "#,##0.00"
                    worksheet[f"{tax_rate_col}{row_idx}"].number_format = "0.00"

        # Inject SUMIFS formulas into Year-wise Tax Summary (after Sales History is written)
        if year_tax_data:
            sales_col_map = _get_sales_history_col_map(writer)
            _build_tax_summary_formulas(
                writer, year_tax_df, {"fy_col": "FY"}, sales_col_map, data_row_mapping=tax_summary_data_row_mapping
            )

            # Add Amount (INR) formulas to Year-wise Tax Summary (data rows only)
            if OPENPYXL_AVAILABLE:
                ws_tax = writer.sheets["Year-wise Tax Summary"]
                ts_col_indices = {cell.value: cell.column for cell in ws_tax[1]}
                amt_usd_idx = ts_col_indices.get("Amount ($)")
                er_idx = ts_col_indices.get("Exchange Rate (USD-INR)")
                amt_inr_idx = ts_col_indices.get("Amount (INR)")
                if amt_usd_idx and er_idx and amt_inr_idx:
                    amt_usd_letter = get_column_letter(amt_usd_idx)
                    er_letter = get_column_letter(er_idx)
                    amt_inr_letter = get_column_letter(amt_inr_idx)
                    for row_idx, _ in tax_summary_data_row_mapping:
                        ws_tax[f"{amt_inr_letter}{row_idx}"] = (
                            f'=IF(AND(ISNUMBER({amt_usd_letter}{row_idx}), ISNUMBER({er_letter}{row_idx}), {er_letter}{row_idx}<>0), {amt_usd_letter}{row_idx} * {er_letter}{row_idx}, "")'
                        )
                        ws_tax[f"{amt_inr_letter}{row_idx}"].number_format = "#,##0.00"

        # Create tax withholding sheet
        tax_data = []
        for grant_id, grant in grants.items():
            for tax in grant["tax_withholdings"]:
                tax_row = {
                    "Grant ID": grant["grant_id"],
                    "Grant Type": grant.get("grant_type", "RSU"),
                    "Symbol": grant["symbol"],
                    "Grant Date": grant["grant_date_str"],
                    "Tax Description": tax["tax_description"],
                    "Tax Rate (%)": tax["tax_rate"],
                    "Withholding Amount ($)": tax["withholding_amount"],
                }
                tax_data.append(tax_row)

        if tax_data:
            tax_df = pd.DataFrame(tax_data)
            # Sort by Grant Date
            tax_df["Grant Date Parsed"] = pd.to_datetime(tax_df["Grant Date"], errors="coerce")
            tax_df = tax_df.sort_values("Grant Date Parsed", ascending=False)
            tax_df = tax_df.drop("Grant Date Parsed", axis=1)
            tax_df.to_excel(writer, sheet_name="Tax Withholdings", index=False)

        # Create Schedule FA Table A3 sheet (handles its own formatting internally)
        _write_schedule_fa_table_a3(writer, grants)

        # --- Apply formatting to all sheets ---
        if OPENPYXL_AVAILABLE:
            _format_worksheet(writer.sheets["Grant Summary"])
            if year_tax_data:
                _format_worksheet(writer.sheets["Year-wise Tax Summary"], skip_rows=tax_summary_subtotal_rows)
            if vesting_data:
                _format_worksheet(writer.sheets["Vesting Schedule"])
            if sales_data:
                _format_worksheet(writer.sheets["Sales History"])
            if tax_data:
                _format_worksheet(writer.sheets["Tax Withholdings"])

    print(f"Summary created successfully: {output_file}")

    # Check for validation issues
    issues_df = summary_df[summary_df["Validation Status"] != "OK"]
    if not issues_df.empty:
        print(f"\n[WARNING]  Validation issues found in {len(issues_df)} grants:")
        for idx, row in issues_df.iterrows():
            print(f"  - {row['Grant ID']}: {row['Validation Status']}")
    else:
        print("\n[OK] All grants passed validation checks!")

    return summary_df


def process_rsu_tracker(input_file, output_file, symbol_for_price="PTC"):
    """
    Process RSU tracker Excel file and generate structured summary.
    [DEPRECATED: Use process_benefit_history() for new multi-sheet format]

    Parameters:
    -----------
    input_file : str
        Path to input Excel file
    output_file : str
        Path for output Excel file
    symbol_for_price : str
        Stock ticker symbol for historical price lookup (default: 'PTC')
    """

    print(f"Reading file: {input_file}")

    # Read the Excel file - try new format first
    try:
        espp_df = pd.read_excel(input_file, sheet_name="ESPP")
    except Exception:
        espp_df = None

    try:
        rs_df = pd.read_excel(input_file, sheet_name="Restricted Stock")
    except Exception:
        rs_df = None

    # If new format not found, try single sheet
    if rs_df is None and espp_df is None:
        df = pd.read_excel(input_file)
    else:
        # Redirect to new function
        return process_benefit_history(input_file, output_file, symbol_for_price)

    # Standardize column names (strip whitespace)
    df.columns = df.columns.str.strip()

    # Remove completely empty rows
    df = df.dropna(how="all")

    # Reset index for easier processing
    df = df.reset_index(drop=True)

    # Dictionary to store grant information
    grants = {}
    current_grant = None
    grant_counter = 0

    # Load overrides for this legacy single-sheet path
    overrides = load_sale_price_overrides()
    _sale_seq_counter = {}  # {(grant_id, sale_date_iso): next_seq}

    print("Processing data...")

    # Process each row
    for idx, row in df.iterrows():
        record_type = str(row["Record Type"]).strip() if pd.notna(row.get("Record Type")) else ""

        # Handle Grant records
        if record_type == "Grant":
            grant_counter += 1
            symbol = str(row["Symbol"]).strip() if pd.notna(row.get("Symbol")) else ""
            grant_date_str = str(row["Grant Date"]).strip() if pd.notna(row.get("Grant Date")) else ""

            # Create unique grant ID (date + counter for duplicates)
            grant_id = f"{grant_date_str}_{grant_counter}"

            # Parse grant date
            grant_date = parse_date(grant_date_str)

            # Get grant date stock price for capital gains calculation
            grant_price = None
            if YFINANCE_AVAILABLE and symbol:
                grant_price = get_stock_price(symbol, grant_date_str)

            # Initialize grant dictionary
            current_grant = {
                "grant_id": grant_id,
                "symbol": symbol,
                "grant_date": grant_date,
                "grant_date_str": grant_date_str,
                "grant_price": grant_price,
                "granted_qty": float(row["Granted Qty."]) if pd.notna(row.get("Granted Qty.")) else 0,
                "withheld_qty": float(row["Withheld Qty."]) if pd.notna(row.get("Withheld Qty.")) else 0,
                "vested_qty": float(row["Vested Qty."]) if pd.notna(row.get("Vested Qty.")) else 0,
                "sellable_qty": float(row["Sellable Qty."]) if pd.notna(row.get("Sellable Qty.")) else 0,
                "unvested_qty": float(row["Unvested Qty."]) if pd.notna(row.get("Unvested Qty.")) else 0,
                "released_qty": float(row["Released Qty"]) if pd.notna(row.get("Released Qty")) else 0,
                "est_market_value": float(row["Est. Market Value"]) if pd.notna(row.get("Est. Market Value")) else 0,
                "events": [],  # List of events (vest, release, sell)
                "vest_tranches": [],  # Per-vest-tranche data for cost basis
                "vest_schedules": [],  # List of vest schedules
                "tax_withholdings": [],  # List of tax withholdings
                "sales": [],  # List of sales
                "capital_gains_tax": [],  # List of capital gains taxes
                "total_tax_withheld": 0,
                "total_capital_gains_tax": 0,
                "total_sold_qty": 0,
                "total_sale_proceeds": 0,
                "sale_dates": [],
                "validation_issues": [],
            }

            grants[grant_id] = current_grant

        # Handle Event records (grant, vest, release, sell)
        elif record_type == "Event" and current_grant is not None:
            event_date_str = str(row["Date"]).strip() if pd.notna(row.get("Date")) else ""
            event_type = str(row["Event Type"]).strip() if pd.notna(row.get("Event Type")) else ""
            qty_or_amount = float(row["Qty. or Amount"]) if pd.notna(row.get("Qty. or Amount")) else 0

            event_date = parse_date(event_date_str)

            event_info = {"date": event_date, "date_str": event_date_str, "type": event_type, "quantity": qty_or_amount}

            current_grant["events"].append(event_info)

            # Track vest tranches for RSU cost basis (vest date = acquisition date)
            if "vested" in event_type.lower():
                vest_price = get_stock_price(symbol_for_price, event_date_str) if YFINANCE_AVAILABLE else None
                current_grant["vest_tranches"].append(
                    {
                        "vest_date": event_date,
                        "vest_date_str": event_date_str,
                        "quantity": qty_or_amount,
                        "vest_price": vest_price,
                    }
                )

            # Track sales separately
            if "sold" in event_type.lower():
                # Determine sequence number for this sale (handles same-day multi-sales)
                _parsed_sale = parse_date(event_date_str)
                _sale_date_iso = _parsed_sale.strftime("%Y-%m-%d") if _parsed_sale else event_date_str
                _seq_key = (current_grant["grant_id"], _sale_date_iso)
                sale_seq = _sale_seq_counter.get(_seq_key, 0) + 1
                _sale_seq_counter[_seq_key] = sale_seq

                sale_price, price_source, _write_new = resolve_sale_price(
                    current_grant["grant_id"], event_date_str, row, symbol_for_price, overrides, sale_seq
                )
                if _write_new:
                    overrides[(current_grant["grant_id"], _sale_date_iso, sale_seq)] = {
                        "sale_price_usd": sale_price,
                        "sale_quantity": qty_or_amount,
                        "source": price_source,
                        "notes": "",
                    }

                # Get exchange rate on sale date
                exchange_rate = None
                if YFINANCE_AVAILABLE:
                    exchange_rate = get_exchange_rate(event_date_str)

                # Match sale to vest tranche (most recent vest before or on sale date)
                matched_vest = None
                for vt in reversed(current_grant["vest_tranches"]):
                    if vt["vest_date"] and event_date and vt["vest_date"] <= event_date:
                        matched_vest = vt
                        break

                # Use vest date/price as acquisition date/cost basis for RSUs
                if matched_vest:
                    acquisition_date = matched_vest["vest_date"]
                    cost_basis_price = matched_vest["vest_price"]
                else:
                    # Fallback to grant date if no vest tranche found
                    acquisition_date = current_grant["grant_date"]
                    cost_basis_price = current_grant["grant_price"]

                # Calculate capital gains tax based on holding period
                capital_gain = 0
                capital_gains_tax = 0
                tax_rate = 0
                tax_type = "N/A"

                if sale_price is not None and cost_basis_price is not None:
                    capital_gain = (sale_price - cost_basis_price) * qty_or_amount

                    # Determine tax rate based on holding period from vest date
                    tax_rate, tax_type = get_capital_gains_tax_rate(acquisition_date, event_date)

                    if tax_rate is not None:
                        capital_gains_tax = capital_gain * tax_rate
                        current_grant["total_capital_gains_tax"] += capital_gains_tax

                        holding_days = (event_date - acquisition_date).days if acquisition_date else 0

                        # Track capital gain tax separately
                        current_grant["capital_gains_tax"].append(
                            {
                                "date": event_date,
                                "date_str": event_date_str,
                                "grant_price": cost_basis_price,
                                "sale_price": sale_price,
                                "quantity": qty_or_amount,
                                "capital_gain": capital_gain,
                                "holding_days": holding_days,
                                "tax_type": tax_type,
                                "tax_rate": tax_rate,
                                "tax_amount": capital_gains_tax,
                            }
                        )

                holding_days = (event_date - acquisition_date).days if acquisition_date and event_date else 0

                sale_info = {
                    "date": event_date,
                    "date_str": event_date_str,
                    "quantity": qty_or_amount,
                    "price": sale_price,
                    "price_source": price_source,
                    "grant_price": cost_basis_price,  # FMV on vest date
                    "acquisition_date": acquisition_date,  # Vest date for holding period
                    "capital_gain": capital_gain,
                    "capital_gains_tax": capital_gains_tax,
                    "holding_days": holding_days,
                    "tax_type": tax_type,
                    "tax_rate": tax_rate,
                    "exchange_rate": exchange_rate,
                }
                current_grant["sales"].append(sale_info)
                current_grant["total_sold_qty"] += qty_or_amount
                current_grant["sale_dates"].append(event_date_str)

        # Handle Vest Schedule records
        elif record_type == "Vest Schedule" and current_grant is not None:
            vest_date_str = str(row["Vest Date"]).strip() if pd.notna(row.get("Vest Date")) else ""
            vested_qty = float(row["Vested Qty."]) if pd.notna(row.get("Vested Qty.")) else 0
            released_qty = float(row["Released Qty"]) if pd.notna(row.get("Released Qty")) else 0
            vest_period = str(row["Vest Period"]).strip() if pd.notna(row.get("Vest Period")) else ""

            vest_date = parse_date(vest_date_str)

            vest_schedule = {
                "vest_date": vest_date,
                "vest_date_str": vest_date_str,
                "vested_qty": vested_qty,
                "released_qty": released_qty,
                "vest_period": vest_period,
                "is_future": vest_date > datetime.now() if vest_date else False,
            }

            current_grant["vest_schedules"].append(vest_schedule)

        # Handle Tax Withholding records
        elif record_type == "Tax Withholding" and current_grant is not None:
            withholding_date_str = str(row["Date"]).strip() if pd.notna(row.get("Date")) else ""
            tax_rate = parse_percentage(row["Effective Tax Rate"]) if pd.notna(row.get("Effective Tax Rate")) else 0
            withholding_amount = float(row["Withholding Amount"]) if pd.notna(row.get("Withholding Amount")) else 0
            tax_description = str(row["Tax Description"]).strip() if pd.notna(row.get("Tax Description")) else ""

            # Only include non-zero tax rate withholdings
            if tax_rate > 0:
                withholding_date = parse_date(withholding_date_str)

                # Get exchange rate on withholding date
                exchange_rate = None
                if YFINANCE_AVAILABLE and withholding_date_str:
                    exchange_rate = get_exchange_rate(withholding_date_str)

                tax_info = {
                    "date": withholding_date,
                    "date_str": withholding_date_str,
                    "tax_rate": tax_rate,
                    "withholding_amount": withholding_amount,
                    "tax_description": tax_description,
                    "exchange_rate": exchange_rate,
                }

                current_grant["tax_withholdings"].append(tax_info)
                current_grant["total_tax_withheld"] += withholding_amount

    # Persist any newly discovered sale prices for the legacy single-sheet path
    save_sale_price_overrides(overrides)

    print(f"Found {len(grants)} grants")

    # Process and validate each grant
    summary_data = []

    for grant_id, grant in grants.items():
        # Calculate derived values
        total_released = sum(event["quantity"] for event in grant["events"] if "released" in event["type"].lower())

        # Calculate future vesting from schedules
        future_vesting_qty = sum(
            schedule["vested_qty"] for schedule in grant["vest_schedules"] if schedule["is_future"]
        )

        # Calculate next vest date
        future_vest_dates = [
            schedule["vest_date"]
            for schedule in grant["vest_schedules"]
            if schedule["is_future"] and schedule["vest_date"]
        ]
        next_vest_date = min(future_vest_dates) if future_vest_dates else None

        # Calculate sellable quantity
        # RSU: Released already accounts for ~30% tax withholding at vesting
        calculated_sellable = total_released - grant["total_sold_qty"]

        # Calculate unvested quantity (alternative calculation)
        calculated_unvested = grant["granted_qty"] - grant["vested_qty"]

        # Validation checks
        validation_issues = []

        # Check 1: Granted = Vested + Unvested
        if abs(grant["granted_qty"] - (grant["vested_qty"] + grant["unvested_qty"])) > 0.01:
            validation_issues.append(
                f"Granted ({grant['granted_qty']}) ≠ Vested ({grant['vested_qty']}) + Unvested ({grant['unvested_qty']})"
            )

        # Check 2: Sellable Qty matches calculation
        if abs(grant["sellable_qty"] - calculated_sellable) > 0.01:
            validation_issues.append(
                f"Sellable Qty mismatch: Stored={grant['sellable_qty']}, Calculated={calculated_sellable}"
            )

        # Check 3: Unvested Qty matches calculation
        if abs(grant["unvested_qty"] - calculated_unvested) > 0.01:
            validation_issues.append(
                f"Unvested Qty mismatch: Stored={grant['unvested_qty']}, Calculated={calculated_unvested}"
            )

        # Note: Removed checks 4 & 5 comparing Events vs Schedules
        # These can legitimately differ due to:
        # - Tax withholding reducing actual vesting
        # - Events and Schedules coming from different data sources
        # - Planned schedules vs actual outcomes

        # Format sale dates
        sale_dates_str = "; ".join(sorted(set(grant["sale_dates"]))) if grant["sale_dates"] else "None"

        # Format next vest date
        next_vest_str = next_vest_date.strftime("%Y-%m-%d") if next_vest_date else "N/A"

        # Format validation issues
        validation_str = " | ".join(validation_issues) if validation_issues else "OK"

        # Prepare summary row
        summary_row = {
            "Grant Type": grant.get("grant_type", "RSU"),  # RSU or ESPP
            "Grant ID": grant["grant_id"],
            "Symbol": grant["symbol"],
            "Grant Date": grant["grant_date_str"],
            "Units": grant["granted_qty"],
            "Vested to Date": grant["vested_qty"],
            "Withheld for Taxes": grant["withheld_qty"],
            "Released to Account": total_released,
            "Tax Withheld ($)": grant["total_tax_withheld"],
            "Sold": grant["total_sold_qty"],
            "Sale Dates": sale_dates_str,
            "Sellable": grant["sellable_qty"],
            "Calc Sellable": calculated_sellable,
            "Unvested": grant["unvested_qty"],
            "Calc Unvested": calculated_unvested,
            "Future Vesting (from schedules)": future_vesting_qty,
            "Next Vest Date": next_vest_str,
            "Estimated Market Value ($)": grant["est_market_value"],
            "Validation Status": validation_str,
            "# of Sales": len(grant["sales"]),
            "# of Vest Schedules": len(grant["vest_schedules"]),
            "# of Tax Withholdings": len(grant["tax_withholdings"]),
        }

        summary_data.append(summary_row)

    # Create summary DataFrame
    summary_df = pd.DataFrame(summary_data)

    # Sort by Grant Date
    summary_df["Grant Date Parsed"] = pd.to_datetime(summary_df["Grant Date"], errors="coerce")
    summary_df = summary_df.sort_values("Grant Date Parsed", ascending=False)
    summary_df = summary_df.drop("Grant Date Parsed", axis=1)

    # Create additional sheets for detailed views
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Main summary sheet
        summary_df.to_excel(writer, sheet_name="Grant Summary", index=False)

        # Create year-wise tax summary sheet
        year_tax_data = []
        tax_summary_subtotal_rows = set()
        tax_summary_data_row_mapping = []

        # Add withholding taxes
        for grant_id, grant in grants.items():
            for tax in grant["tax_withholdings"]:
                fy = get_financial_year(tax["date"]) if tax["date"] else get_financial_year(grant["grant_date"])
                if fy:
                    # Use exchange rate that was fetched during data processing, default to None if not available
                    exchange_rate = tax.get("exchange_rate", None)

                    year_tax_row = {
                        "Financial Year": fy,
                        "Tax Date": tax["date_str"],
                        "Grant ID": grant["grant_id"],
                        "Symbol": grant["symbol"],
                        "Tax Type": "Withholding Tax",
                        "Tax Description": tax["tax_description"],
                        "Rate (%)": tax["tax_rate"],
                        "Amount ($)": tax["withholding_amount"],
                        "Exchange Rate (USD-INR)": exchange_rate,
                        "Amount (INR)": None,  # Will be formula
                        "_tax_type_base": None,
                        "_is_capital_gains": False,
                    }
                    year_tax_data.append(year_tax_row)

        # Add capital gains taxes
        for grant_id, grant in grants.items():
            for cg_tax in grant["capital_gains_tax"]:
                fy = get_financial_year(cg_tax["date"])
                if fy:
                    holding_days_display = (
                        f"{cg_tax['holding_days']} days"
                        if cg_tax["holding_days"] <= 365
                        else f"{cg_tax['holding_days'] // 365} years {cg_tax['holding_days'] % 365} days"
                    )
                    # Get exchange rate from the sale that was already fetched
                    exchange_rate = None
                    for sale in grant["sales"]:
                        if sale["date_str"] == cg_tax["date_str"]:
                            exchange_rate = sale["exchange_rate"]
                            break

                    year_tax_row = {
                        "Financial Year": fy,
                        "Tax Date": cg_tax["date_str"],
                        "Grant ID": grant["grant_id"],
                        "Symbol": grant["symbol"],
                        "Tax Type": f"{cg_tax['tax_type']} (Holding: {holding_days_display})",
                        "Tax Description": f"Sale @ ${cg_tax['sale_price']:.2f} (Cost Basis: ${cg_tax['grant_price']:.2f})",
                        "Rate (%)": cg_tax["tax_rate"] * 100,
                        "Amount ($)": cg_tax["tax_amount"],
                        "Exchange Rate (USD-INR)": exchange_rate,
                        "Amount (INR)": None,  # Will be formula
                        "_tax_type_base": cg_tax["tax_type"],
                        "_is_capital_gains": True,
                    }
                    year_tax_data.append(year_tax_row)

        if year_tax_data:
            year_tax_df = pd.DataFrame(year_tax_data)
            # Sort by Financial Year (descending), then by tax amount
            year_tax_df = year_tax_df.sort_values(["Financial Year", "Amount ($)"], ascending=[False, False])

            # Write with FY subtotal rows
            display_cols = [
                "Financial Year",
                "Tax Date",
                "Grant ID",
                "Symbol",
                "Tax Type",
                "Tax Description",
                "Rate (%)",
                "Amount ($)",
                "Exchange Rate (USD-INR)",
                "Amount (INR)",
            ]
            tax_summary_data_row_mapping, tax_summary_subtotal_rows = _write_tax_summary_with_subtotals(
                writer, year_tax_df, "Financial Year", display_cols
            )

        # Create detailed vesting schedule sheet
        vesting_data = []
        for grant_id, grant in grants.items():
            # Calculate expected shares per future vest from grant's unvested qty
            num_future_vests = sum(1 for s in grant["vest_schedules"] if s["is_future"])
            expected_per_vest = (grant["unvested_qty"] / num_future_vests) if num_future_vests > 0 else 0

            for schedule in grant["vest_schedules"]:
                vesting_row = {
                    "Grant ID": grant["grant_id"],
                    "Symbol": grant["symbol"],
                    "Grant Date": grant["grant_date_str"],
                    "Vest Date": schedule["vest_date"],  # datetime for formulas
                    "Vest Period": schedule["vest_period"],
                    "Vested Quantity": schedule["vested_qty"],
                    "Released Quantity": schedule["released_qty"],
                    "Is Future Vesting": "Yes" if schedule["is_future"] else "No",
                    "Days to Vesting": None,  # Formula placeholder
                    "Future Vesting Qty": expected_per_vest if schedule["is_future"] else None,
                }
                vesting_data.append(vesting_row)

        if vesting_data:
            vesting_df = pd.DataFrame(vesting_data)
            # Sort by Vest Date descending (future dates first)
            vesting_df = vesting_df.sort_values("Vest Date", ascending=False)
            vesting_df.to_excel(writer, sheet_name="Vesting Schedule", index=False)

            if OPENPYXL_AVAILABLE:
                ws_vest = writer.sheets["Vesting Schedule"]
                vest_col_indices = {cell.value: cell.column for cell in ws_vest[1]}
                vest_date_col_idx = vest_col_indices.get("Vest Date")
                days_col_idx = vest_col_indices.get("Days to Vesting")

                if vest_date_col_idx:
                    vest_date_letter = get_column_letter(vest_date_col_idx)
                    for row_idx in range(2, len(vesting_data) + 2):
                        ws_vest.cell(row=row_idx, column=vest_date_col_idx).number_format = "YYYY-MM-DD"

                if days_col_idx and vest_date_col_idx:
                    days_letter = get_column_letter(days_col_idx)
                    vest_date_letter = get_column_letter(vest_date_col_idx)
                    for row_idx in range(2, len(vesting_data) + 2):
                        ws_vest[f"{days_letter}{row_idx}"] = (
                            f'=IF({vest_date_letter}{row_idx}>TODAY(), {vest_date_letter}{row_idx}-TODAY(), "")'
                        )
                        ws_vest[f"{days_letter}{row_idx}"].number_format = "0"

        # Create sales history sheet
        sales_data = []
        for grant_id, grant in grants.items():
            for sale in grant["sales"]:
                # Use holding_days from sale (based on vest date for RSUs)
                holding_days = sale.get(
                    "holding_days",
                    (sale["date"] - grant["grant_date"]).days if sale["date"] and grant["grant_date"] else 0,
                )
                holding_display = f"{holding_days} days"
                if holding_days > 365:
                    years = holding_days // 365
                    days = holding_days % 365
                    holding_display = f"{years}y {days}d" if days > 0 else f"{years}y"

                # Store numeric values for formulas
                tax_rate_pct = (sale["tax_rate"] * 100) if sale["tax_rate"] else None

                # Exchange rate display
                exchange_rate_display = sale["exchange_rate"] if sale["exchange_rate"] is not None else 0

                sales_row = {
                    "Grant ID": grant["grant_id"],
                    "Symbol": grant["symbol"],
                    "Grant Date": grant["grant_date_str"],
                    "Sale Date": sale["date"],  # datetime for SUMIFS date comparisons
                    "Holding Period": holding_display,
                    "Quantity Sold": sale["quantity"],
                    "Grant Price ($)": sale["grant_price"],
                    "Sale Price ($)": sale["price"],
                    "Capital Gain ($)": None,  # Will be calculated by formula
                    "Tax Rate (%)": tax_rate_pct,
                    "Tax Type": sale["tax_type"],
                    "Capital Gains Tax ($)": None,  # Will be calculated by formula
                    "Estimated Proceeds ($)": None,  # Will be calculated by formula
                    "Exchange Rate (USD-INR)": exchange_rate_display,
                    "Estimated Proceeds (INR)": None,  # Will be replaced with formula
                    "Capital Gain (INR)": None,  # Will be calculated by formula
                    "Capital Gains Tax (INR)": None,  # Will be calculated by formula
                }
                sales_data.append(sales_row)

        if sales_data:
            sales_df = pd.DataFrame(sales_data)
            # Sort by Sale Date (latest first) — Sale Date is already datetime
            sales_df = sales_df.sort_values("Sale Date", ascending=False)

            sales_df.to_excel(writer, sheet_name="Sales History", index=False)

            # Add formulas and formatting for calculated columns
            if OPENPYXL_AVAILABLE:
                worksheet = writer.sheets["Sales History"]

                # Find the column indices
                col_indices = {}
                for col_idx, col_cell in enumerate(worksheet[1], 1):
                    col_indices[col_cell.value] = col_idx

                # Format Sale Date column as date
                sale_date_col_idx = col_indices.get("Sale Date", None)
                if sale_date_col_idx:
                    for row_idx in range(2, len(sales_data) + 2):
                        worksheet.cell(row=row_idx, column=sale_date_col_idx).number_format = "YYYY-MM-DD"

                # Add formulas for calculated columns (starting from row 2)
                for row_idx in range(2, len(sales_data) + 2):
                    # Get column letters for reference
                    qty_col = get_column_letter(col_indices.get("Quantity Sold", 1))
                    grant_price_col = get_column_letter(col_indices.get("Grant Price ($)", 1))
                    sale_price_col = get_column_letter(col_indices.get("Sale Price ($)", 1))
                    tax_rate_col = get_column_letter(col_indices.get("Tax Rate (%)", 1))
                    cap_gain_col = get_column_letter(col_indices.get("Capital Gain ($)", 1))
                    cap_gain_tax_col = get_column_letter(col_indices.get("Capital Gains Tax ($)", 1))
                    proceeds_col = get_column_letter(col_indices.get("Estimated Proceeds ($)", 1))
                    exchange_col = get_column_letter(col_indices.get("Exchange Rate (USD-INR)", 1))
                    proceeds_inr_col = get_column_letter(col_indices.get("Estimated Proceeds (INR)", 1))
                    cap_gain_inr_col = get_column_letter(col_indices.get("Capital Gain (INR)", 1))
                    cap_gain_tax_inr_col = get_column_letter(col_indices.get("Capital Gains Tax (INR)", 1))

                    # Capital Gain ($) = (Sale Price - Grant Price) * Quantity
                    worksheet[f"{cap_gain_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({sale_price_col}{row_idx}), ISNUMBER({grant_price_col}{row_idx}), ISNUMBER({qty_col}{row_idx})), ({sale_price_col}{row_idx} - {grant_price_col}{row_idx}) * {qty_col}{row_idx}, "")'
                    )

                    # Capital Gains Tax ($) = Capital Gain * Tax Rate / 100
                    worksheet[f"{cap_gain_tax_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({cap_gain_col}{row_idx}), ISNUMBER({tax_rate_col}{row_idx})), {cap_gain_col}{row_idx} * {tax_rate_col}{row_idx} / 100, "")'
                    )

                    # Estimated Proceeds ($) = Quantity * Sale Price
                    worksheet[f"{proceeds_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({qty_col}{row_idx}), ISNUMBER({sale_price_col}{row_idx})), {qty_col}{row_idx} * {sale_price_col}{row_idx}, "")'
                    )

                    # Estimated Proceeds (INR) = Estimated Proceeds ($) * Exchange Rate
                    worksheet[f"{proceeds_inr_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({proceeds_col}{row_idx}), ISNUMBER({exchange_col}{row_idx}), {exchange_col}{row_idx} <> 0), {proceeds_col}{row_idx} * {exchange_col}{row_idx}, "")'
                    )

                    # Capital Gain (INR) = Capital Gain ($) * Exchange Rate
                    worksheet[f"{cap_gain_inr_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({cap_gain_col}{row_idx}), ISNUMBER({exchange_col}{row_idx}), {exchange_col}{row_idx}<>0), {cap_gain_col}{row_idx} * {exchange_col}{row_idx}, "")'
                    )

                    # Capital Gains Tax (INR) = Capital Gains Tax ($) * Exchange Rate
                    worksheet[f"{cap_gain_tax_inr_col}{row_idx}"] = (
                        f'=IF(AND(ISNUMBER({cap_gain_tax_col}{row_idx}), ISNUMBER({exchange_col}{row_idx}), {exchange_col}{row_idx}<>0), {cap_gain_tax_col}{row_idx} * {exchange_col}{row_idx}, "")'
                    )

                    # Apply currency formatting
                    for col in [
                        cap_gain_col,
                        sale_price_col,
                        grant_price_col,
                        cap_gain_tax_col,
                        proceeds_col,
                        proceeds_inr_col,
                    ]:
                        worksheet[f"{col}{row_idx}"].number_format = "$#,##0.00"
                    for col in [cap_gain_inr_col, cap_gain_tax_inr_col]:
                        worksheet[f"{col}{row_idx}"].number_format = "#,##0.00"
                    worksheet[f"{tax_rate_col}{row_idx}"].number_format = "0.00"

        # Inject SUMIFS formulas into Year-wise Tax Summary (after Sales History is written)
        if year_tax_data:
            sales_col_map = _get_sales_history_col_map(writer)
            _build_tax_summary_formulas(
                writer,
                year_tax_df,
                {"fy_col": "Financial Year"},
                sales_col_map,
                data_row_mapping=tax_summary_data_row_mapping,
            )

            # Add Amount (INR) formulas to Year-wise Tax Summary (data rows only)
            if OPENPYXL_AVAILABLE:
                ws_tax = writer.sheets["Year-wise Tax Summary"]
                ts_col_indices = {cell.value: cell.column for cell in ws_tax[1]}
                amt_usd_idx = ts_col_indices.get("Amount ($)")
                er_idx = ts_col_indices.get("Exchange Rate (USD-INR)")
                amt_inr_idx = ts_col_indices.get("Amount (INR)")
                if amt_usd_idx and er_idx and amt_inr_idx:
                    amt_usd_letter = get_column_letter(amt_usd_idx)
                    er_letter = get_column_letter(er_idx)
                    amt_inr_letter = get_column_letter(amt_inr_idx)
                    for row_idx, _ in tax_summary_data_row_mapping:
                        ws_tax[f"{amt_inr_letter}{row_idx}"] = (
                            f'=IF(AND(ISNUMBER({amt_usd_letter}{row_idx}), ISNUMBER({er_letter}{row_idx}), {er_letter}{row_idx}<>0), {amt_usd_letter}{row_idx} * {er_letter}{row_idx}, "")'
                        )
                        ws_tax[f"{amt_inr_letter}{row_idx}"].number_format = "#,##0.00"

        # Create tax withholding sheet
        tax_data = []
        for grant_id, grant in grants.items():
            for tax in grant["tax_withholdings"]:
                tax_row = {
                    "Grant ID": grant["grant_id"],
                    "Symbol": grant["symbol"],
                    "Grant Date": grant["grant_date_str"],
                    "Tax Description": tax["tax_description"],
                    "Tax Rate (%)": tax["tax_rate"],
                    "Withholding Amount ($)": tax["withholding_amount"],
                }
                tax_data.append(tax_row)

        if tax_data:
            tax_df = pd.DataFrame(tax_data)
            # Sort by Grant Date
            tax_df["Grant Date Parsed"] = pd.to_datetime(tax_df["Grant Date"], errors="coerce")
            tax_df = tax_df.sort_values("Grant Date Parsed", ascending=False)
            tax_df = tax_df.drop("Grant Date Parsed", axis=1)
            tax_df.to_excel(writer, sheet_name="Tax Withholdings", index=False)

        # --- Apply formatting to all sheets ---
        if OPENPYXL_AVAILABLE:
            _format_worksheet(writer.sheets["Grant Summary"])
            if year_tax_data:
                _format_worksheet(writer.sheets["Year-wise Tax Summary"], skip_rows=tax_summary_subtotal_rows)
            if vesting_data:
                _format_worksheet(writer.sheets["Vesting Schedule"])
            if sales_data:
                _format_worksheet(writer.sheets["Sales History"])
            if tax_data:
                _format_worksheet(writer.sheets["Tax Withholdings"])

    print(f"Summary created successfully: {output_file}")

    # Check for validation issues
    issues_df = summary_df[summary_df["Validation Status"] != "OK"]
    if not issues_df.empty:
        print(f"\n[WARNING]  Validation issues found in {len(issues_df)} grants:")
        for idx, row in issues_df.iterrows():
            print(f"  - {row['Grant ID']}: {row['Validation Status']}")
    else:
        print("\n[OK] All grants passed validation checks!")

    return summary_df


def main():
    config = configparser.ConfigParser()
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vestwise.ini")
    config.read(config_path)

    input_file = config.get("paths", "input_file", fallback="BenefitHistory.xlsx")
    out_template = config.get("paths", "output_file_template", fallback="{timestamp}_rsu_summary.xlsx")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = out_template.format(timestamp=timestamp)

    # Process the benefit history (RSU and ESPP)
    process_rsu_tracker(input_file, output_file)

    # Display sample of the summary
    # print("\nSample of the summary (first 5 grants):")
    # print(summary_df.head().to_string())


if __name__ == "__main__":
    main()
